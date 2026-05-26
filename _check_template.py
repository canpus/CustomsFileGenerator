# -*- coding: utf-8 -*-
"""临时脚本：检查装箱单模板结构."""
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("templates/template_packing.xlsx")
ws = wb.active
print("Sheet name:", ws.title)
print("Max row:", ws.max_row, "Max col:", ws.max_column)
print()

for row in range(1, min(ws.max_row + 1, 75)):
    vals = []
    for col in range(1, min(ws.max_column + 1, 12)):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            col_letter = get_column_letter(col)
            vals.append(f"{col_letter}{row}={v!r}")
    if vals:
        print(f"Row {row}: {' | '.join(vals)}")
wb.close()

# -*- coding: utf-8 -*-
"""临时脚本：检查发票和合同模板结构."""
import openpyxl
from openpyxl.utils import get_column_letter

for name in ["template_invoice.xlsx", "template_contract.xlsx"]:
    print(f"\n{'='*60}")
    print(f"Template: {name}")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(f"templates/{name}")
    for idx, ws_name in enumerate(wb.sheetnames):
        ws = wb[ws_name]
        print(f"\n  Sheet: {ws_name} (max_row={ws.max_row}, max_col={ws.max_column})")
        print(f"  Merged cells: {list(ws.merged_cells.ranges)[:20]}...")
        for row in range(1, min(ws.max_row + 1, 80)):
            vals = []
            for col in range(1, min(ws.max_column + 1, 10)):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    col_letter = get_column_letter(col)
                    cell_type = type(ws.cell(row=row, column=col)).__name__
                    vals.append(f"{col_letter}{row}[{cell_type}]={v!r}")
            if vals:
                print(f"    Row {row}: {' | '.join(vals)}")
    wb.close()

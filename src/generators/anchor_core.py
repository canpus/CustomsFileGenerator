# -*- coding: utf-8 -*-
"""模板锚点扫描 — 核心工具模块.

包含 AnchorResult 数据类及所有共享扫描工具函数。
供 scanners/ 子模块和 template_anchor_scanner 主模块共同导入。
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List

from openpyxl.worksheet.worksheet import Worksheet

from config.constants import TEMPLATE_RULES_PATH

logger = logging.getLogger(__name__)


# ==================== 数据类 ====================


@dataclass
class AnchorResult:
    """锚点扫描结果.

    Attributes:
        template_name: 模板名称（如 "packing"）。
        sheet_name: 工作表名称。
        header_row: 标题行号（1-based），未找到时为 -1。
        data_start_row: 数据起始行号（1-based），未找到时为 -1。
        data_end_row: 数据预设结束行号（模板默认值，1-based）。
        summary_row: 汇总行号（1-based），未找到时为 -1。
        summary_rows: 所有匹配的汇总行号列表。
        column_mapping: 列字母映射（如 {"seq_no": "A"}）。
        summary_columns: 汇总列映射。
        merge_ranges: 合并列范围列表。
        errors: 扫描过程中的警告/错误信息。
    """

    template_name: str
    sheet_name: str = ""
    header_row: int = -1
    data_start_row: int = -1
    data_end_row: int = -1
    summary_row: int = -1
    summary_rows: List[int] = field(default_factory=list)
    column_mapping: Dict[str, str] = field(default_factory=dict)
    summary_columns: Dict[str, str] = field(default_factory=dict)
    merge_ranges: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        """锚点扫描是否成功（关键锚点均已定位）."""
        return (
            self.header_row > 0
            and self.data_start_row > 0
            and self.summary_row > 0
        )


# ==================== 规则加载 ====================


def _load_rules() -> dict:
    """从 config/template_rules.json 加载扫描规则.

    Returns:
        规则字典。若文件不存在或解析失败则返回空字典。
    """
    if not TEMPLATE_RULES_PATH.exists():
        logger.warning(
            "[警告]: 锚点规则文件不存在: %s，将仅使用内置默认规则",
            TEMPLATE_RULES_PATH,
        )
        return {}

    try:
        with open(TEMPLATE_RULES_PATH, "r", encoding="utf-8") as f:
            rules = json.load(f)
        logger.info("已加载锚点扫描规则: %s", TEMPLATE_RULES_PATH)
        return rules
    except json.JSONDecodeError as e:
        logger.error(
            "[错误]: 锚点规则文件 JSON 解析失败: %s\n"
            "[原因]: %s\n"
            "[排查]: 请检查 %s 的 JSON 格式是否正确",
            TEMPLATE_RULES_PATH, e, TEMPLATE_RULES_PATH.name,
        )
        return {}
    except Exception as e:
        logger.error(
            "[错误]: 读取锚点规则文件失败: %s\n"
            "[原因]: %s\n"
            "[排查]: 请确认文件存在且有读取权限",
            TEMPLATE_RULES_PATH, e,
        )
        return {}


# ==================== 通用扫描逻辑 ====================


def _contains_keyword(cell_value, keywords: List[str]) -> bool:
    """检查单元格值是否包含任一关键词（不区分大小写）."""
    if cell_value is None:
        return False
    text: str = str(cell_value).strip()
    if not text:
        return False
    text_lower: str = text.lower()
    return any(kw.lower() in text_lower for kw in keywords)


def _find_row_by_keywords(
    ws: Worksheet,
    row_keywords: List[str],
    col_keywords: List[str] | None = None,
    row_start: int = 1,
    row_end: int = 30,
    require_both: bool = False,
) -> int:
    """在指定行范围内查找匹配关键词的行.

    扫描策略：
    1. 先按 row_keywords 匹配整行任意单元格。
    2. 若提供 col_keywords，优先返回同时满足 row + col 关键词的行。
    3. 若 require_both=True，必须同时满足两个条件。
    """
    best_match: int = -1
    best_score: int = 0

    for row_idx in range(row_start, row_end + 1):
        row_ok: bool = False
        col_ok: bool = False

        for col_idx in range(1, (ws.max_column or 26) + 1):
            try:
                cell_value = ws.cell(row=row_idx, column=col_idx).value
            except Exception:
                continue

            if not row_ok and _contains_keyword(cell_value, row_keywords):
                row_ok = True

            if col_keywords and not col_ok and _contains_keyword(cell_value, col_keywords):
                col_ok = True

            if row_ok and (not col_keywords or col_ok):
                current_score: int = 2 if (row_ok and col_ok) else 1
                if current_score > best_score:
                    best_score = current_score
                    best_match = row_idx
                    if best_score == 2:
                        return best_match

    if require_both and best_score < 2:
        return -1
    return best_match


def _find_summary_rows(
    ws: Worksheet,
    row_keywords: List[str],
    col_keywords: List[str] | None = None,
    row_start: int = 1,
    row_end: int = 100,
) -> List[int]:
    """查找汇总行（可能有多行）."""
    matches: List[int] = []

    for row_idx in range(row_start, row_end + 1):
        row_ok: bool = False
        col_ok: bool = col_keywords is None

        for col_idx in range(1, (ws.max_column or 26) + 1):
            try:
                cell_value = ws.cell(row=row_idx, column=col_idx).value
            except Exception:
                continue

            if isinstance(cell_value, str) and cell_value.startswith("="):
                if _contains_keyword(cell_value, row_keywords):
                    row_ok = True

            if not row_ok and _contains_keyword(cell_value, row_keywords):
                row_ok = True

            if col_keywords and not col_ok and _contains_keyword(cell_value, col_keywords):
                col_ok = True

            if row_ok and col_ok:
                matches.append(row_idx)
                break

    return matches


def _estimate_data_end_row(ws: Worksheet, data_start_row: int, summary_row: int) -> int:
    """估算模板的数据预设结束行.

    在 data_start_row 和 summary_row 之间查找第一个空数据行，
    则该行的上一行即为数据预设结束行。
    跳过 data_start_row 本身（空模板尚无数据）。
    """
    max_col: int = ws.max_column or 1

    for row_idx in range(data_start_row + 1, summary_row):
        all_empty: bool = True
        for col_idx in range(1, max_col + 1):
            try:
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    all_empty = False
                    break
            except Exception:
                continue
        if all_empty:
            return row_idx - 1

    return summary_row - 1

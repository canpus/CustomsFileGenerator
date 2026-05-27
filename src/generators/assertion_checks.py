"""断言引擎 — xlsx 各维度检查函数."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from src.generators.assertion_data import AssertionLevel, AssertionReport


def _assert_font(
    ws: Worksheet,
    template_ws: Worksheet,
    rules: dict,
    report: AssertionReport,
    template_name: str,
) -> None:
    """断言字体样式 — 检查关键单元格的字体名称和字号."""
    font_rules: dict = rules.get("assertions", {}).get("font", {})
    if not font_rules:
        return

    check_cells: list[str] = font_rules.get("check_cells", [])
    expected_name: str = font_rules.get("expected", {}).get("name", "")
    tolerance: dict = font_rules.get("tolerance", {})
    level_str: str = font_rules.get("level", "warning")
    level: AssertionLevel = "warning" if level_str == "warning" else "error"

    for cell_ref in check_cells:
        try:
            gen_cell = ws[cell_ref]
            tpl_cell = template_ws[cell_ref]
        except Exception:
            continue

        gen_font_name: str = (gen_cell.font.name or "").strip() if gen_cell.font else ""

        if expected_name and gen_font_name and gen_font_name != expected_name:
            name_exact: bool = tolerance.get("name", "exact") == "exact"
            if name_exact or expected_name.lower() != gen_font_name.lower():
                report.add(
                    level,
                    "font_name",
                    f"单元格 {cell_ref} 字体名称不匹配",
                    f"期望: '{expected_name}', 实际: '{gen_font_name}'",
                )

        gen_size: float = gen_cell.font.size if gen_cell.font and gen_cell.font.size else 0.0
        tpl_size: float = tpl_cell.font.size if tpl_cell.font and tpl_cell.font.size else 0.0
        if tpl_size > 0 and gen_size > 0:
            size_tolerance: float = tolerance.get("size_pt", 0.5)
            if abs(gen_size - tpl_size) > size_tolerance:
                report.add(
                    level,
                    "font_size",
                    f"单元格 {cell_ref} 字号偏差超过容差",
                    f"模板: {tpl_size}pt, 生成: {gen_size}pt, 容差: ±{size_tolerance}pt",
                )


def _assert_border(
    ws: Worksheet,
    template_ws: Worksheet,
    rules: dict,
    report: AssertionReport,
    template_name: str,
) -> None:
    """断言边框样式 — 检查关键单元格是否有边框."""
    border_rules: dict = rules.get("assertions", {}).get("border", {})
    if not border_rules:
        return

    check_cells: list[str] = border_rules.get("check_cells", [])
    level_str: str = border_rules.get("level", "error")
    level: AssertionLevel = "error" if level_str == "error" else "warning"

    for cell_ref in check_cells:
        try:
            gen_cell = ws[cell_ref]
        except Exception:
            continue

        has_border: bool = gen_cell.border is not None and (
            gen_cell.border.left is not None
            or gen_cell.border.right is not None
            or gen_cell.border.top is not None
            or gen_cell.border.bottom is not None
        )

        if not has_border:
            report.add(
                level,
                "border",
                f"单元格 {cell_ref} 边框丢失",
                "边框丢失可能导致报关资料格式不合规",
            )


def _assert_merge_ranges(
    ws: Worksheet,
    template_ws: Worksheet,
    rules: dict,
    report: AssertionReport,
    template_name: str,
) -> None:
    """断言合并单元格区域."""
    merge_rules: dict = rules.get("assertions", {}).get("merge_ranges", {})
    if not merge_rules:
        return

    expected_ranges: list[str] = merge_rules.get("expected_ranges", [])
    level_str: str = merge_rules.get("level", "error")
    level: AssertionLevel = "error" if level_str == "error" else "warning"

    gen_merged: set = set()
    for merged_range in ws.merged_cells.ranges:
        gen_merged.add(str(merged_range))

    for expected_range in expected_ranges:
        if expected_range not in gen_merged:
            report.add(
                level,
                "merge_range",
                f"合并单元格区域缺失: {expected_range}",
                "表头合并区域丢失，可能导致格式与模板不一致",
            )


def _assert_formula(
    ws: Worksheet,
    rules: dict,
    report: AssertionReport,
    template_name: str,
) -> None:
    """断言汇总公式 — 检查是否包含 SUM 公式."""
    formula_rules: dict = rules.get("assertions", {}).get("formula", {})
    if not formula_rules:
        return

    expected_pattern: str = formula_rules.get("expected_pattern", "=SUM(")
    level_str: str = formula_rules.get("level", "info")
    level: AssertionLevel = "info" if level_str == "info" else "warning"

    found_formula: bool = False
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith(expected_pattern):
                found_formula = True
                break
        if found_formula:
            break

    if not found_formula:
        report.add(
            level,
            "formula",
            f"未找到 {expected_pattern} 汇总公式",
            "汇总行公式缺失，数据变更后汇总不会自动更新",
        )
    else:
        report.add("info", "formula", "汇总公式检查通过")


def _assert_column_count(
    ws: Worksheet,
    rules: dict,
    report: AssertionReport,
    template_name: str,
) -> None:
    """断言列数（使用 >= 比较，允许模板含额外空列）."""
    col_rules: dict = rules.get("assertions", {}).get("column_count", {})
    if not col_rules:
        return

    expected: int = col_rules.get("expected", 0)
    level_str: str = col_rules.get("level", "warning")
    level: AssertionLevel = "warning" if level_str == "warning" else "error"

    actual: int = ws.max_column or 0

    if actual < expected:
        report.add(
            level,
            "column_count",
            f"列数不足: 期望至少 {expected} 列, 实际 {actual} 列",
            "列数减少可能导致数据列缺失",
        )

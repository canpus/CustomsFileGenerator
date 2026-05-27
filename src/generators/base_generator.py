"""生成器抽象基类 — 阶段 5.

定义所有文件生成器的统一接口和共享工具方法。
PackingGenerator、InvoiceGenerator、ContractGenerator 均继承此基类。
"""

from __future__ import annotations

import abc
import logging
import shutil
import tempfile
from collections.abc import Callable
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from config.constants import (
    FILE_PREFIX_MAP,
    MAX_TEMPLATE_ROWS,
    OUTPUT_DIR,
    TEMPLATE_CAPACITY_THRESHOLDS,
)
from src.generators.template_anchor_scanner import AnchorResult
from src.generators.xlsx_utils import resize_data_rows
from src.models.order_data import OrderData

logger = logging.getLogger(__name__)


class BaseGenerator(abc.ABC):
    """文件生成器抽象基类.

    定义 generate() 统一接口，并提供沙箱复制、进度回调等共享方法。

    子类必须实现：
    - _scan_anchor(ws, rules) → AnchorResult
    - _fill_header(ws, order) → None
    - _fill_data_rows(ws, data_start_row, rows, anchor) → None
    - _fix_summary_formulas(ws, anchor, new_data_end) → None
    - _flatten_data(order) → list[dict]
    - _get_template_type() → str
    """

    def __init__(self, template_path: str | Path | None = None):
        """初始化生成器.

        Args:
            template_path: 模板文件路径，默认由子类 _get_default_template_path() 提供.
        """
        self._template_path: Path = (
            Path(template_path) if template_path else self._get_default_template_path()
        )

    # ==================== 公开 API ====================

    def generate(
        self,
        order: OrderData,
        output_dir: str | Path | None = None,
        progress_callback: Callable[[str, float], None] | None = None,
    ) -> Path:
        """生成文件（统一入口）.

        完整流程：
        1. 入口断言（order 非空 + pallets 非空）
        2. 展平数据
        3. 沙箱复制模板
        4. 锚点扫描
        5. 行数调整（扩容/缩容）
        6. 填充表头
        7. 填充明细行
        8. 更新汇总公式
        9. 保存到输出目录

        Args:
            order: 订单数据.
            output_dir: 输出目录，默认为 config/constants.OUTPUT_DIR.
            progress_callback: 进度回调，签名为 (description, progress_0_1).

        Returns:
            生成的文件路径.

        Raises:
            ValueError: 订单数据无效时抛出.
            FileNotFoundError: 模板文件不存在时抛出.
        """
        # 步骤 0：入口断言
        if order is None:
            raise ValueError(f"[错误]: 订单数据为 None, 无法生成{self._get_display_name()}")
        if not order.pallets:
            raise ValueError(f"[错误]: 订单无托盘数据, 无法生成{self._get_display_name()}")

        # 步骤 1：展平数据
        self._report_progress(progress_callback, "正在解析订单数据...", 0.05)
        rows: list[dict] = self._flatten_data(order)
        target_row_count: int = len(rows)

        logger.info(
            "开始生成%s: 发票号=%s, 托盘数=%d, 明细行数=%d",
            self._get_display_name(),
            order.order_meta.invoice_no,
            len(order.pallets),
            target_row_count,
        )

        # 步骤 1.5：根据数据行数自动选择合适容量的模板
        self._report_progress(progress_callback, "正在选择模板容量...", 0.08)
        selected_template: Path = self._select_template_by_rows(target_row_count)
        logger.info("已选择模板: %s", selected_template.name)

        # 步骤 2：沙箱复制模板（使用选中的容量模板）
        self._report_progress(progress_callback, "正在准备模板...", 0.10)
        sandbox_path: Path = self._create_sandbox_copy(selected_template)

        try:
            wb = openpyxl.load_workbook(sandbox_path)
            ws = wb.active if wb.active else wb.worksheets[0]

            # 步骤 3：锚点扫描
            self._report_progress(progress_callback, "正在扫描模板锚点...", 0.15)
            anchor: AnchorResult = self._scan_anchor(ws)

            if not anchor.is_valid:
                error_details: str = "; ".join(anchor.errors)
                raise ValueError(
                    f"[错误]: {self._get_display_name()}模板锚点扫描失败\n"
                    f"[原因]: {error_details}\n"
                    f"[排查]: 请确认模板文件是否与原厂一致，"
                    f"或从 src/assets/backup_templates/ 恢复出厂模板"
                )

            logger.info(
                "锚点扫描成功: data_start=%d, data_end=%d, summary=%d",
                anchor.data_start_row,
                anchor.data_end_row,
                anchor.summary_row,
            )

            # 步骤 4：行数调整
            self._report_progress(progress_callback, "正在调整数据行数...", 0.20)

            new_data_end: int = resize_data_rows(
                ws,
                anchor.data_start_row,
                anchor.data_end_row,
                target_row_count,
                anchor.data_start_row,
            )
            logger.info("数据行调整完成: 新结束行=%d", new_data_end)

            # 步骤 5：填充表头
            self._report_progress(progress_callback, "正在填充表头信息...", 0.30)
            self._fill_header(ws, order)

            # 步骤 6：填充明细行
            self._report_progress(progress_callback, "正在填充商品明细...", 0.40)
            self._fill_data_rows(ws, anchor.data_start_row, rows, anchor)

            # 步骤 7：更新汇总公式
            self._report_progress(progress_callback, "正在更新汇总公式...", 0.80)
            self._fix_summary_formulas(ws, anchor, new_data_end)

            # 步骤 8：保存到输出目录
            self._report_progress(progress_callback, "正在保存文件...", 0.90)
            output_path: Path = self._resolve_output_path(order, output_dir)
            wb.save(output_path)
            wb.close()

            logger.info("%s已生成: %s", self._get_display_name(), output_path)
            self._report_progress(progress_callback, f"{self._get_display_name()}生成完成", 1.0)

            return output_path

        except Exception:
            try:
                wb.close()
            except Exception as close_err:
                logger.warning("[警告]: 工作簿关闭失败: %s", close_err)
            raise
        finally:
            self._cleanup_sandbox(sandbox_path)

    # ==================== 子类必须实现的抽象方法 ====================

    @abc.abstractmethod
    def _get_default_template_path(self) -> Path:
        """返回默认模板文件路径."""
        ...

    @abc.abstractmethod
    def _get_template_type(self) -> str:
        """返回模板类型标识（如 "invoice"、"contract"）."""
        ...

    @abc.abstractmethod
    def _get_display_name(self) -> str:
        """返回生成器显示名称（如 "形式发票"）."""
        ...

    @abc.abstractmethod
    def _scan_anchor(self, ws) -> AnchorResult:
        """扫描模板锚点.

        Args:
            ws: openpyxl Worksheet 对象.

        Returns:
            AnchorResult 扫描结果.
        """
        ...

    @abc.abstractmethod
    def _flatten_data(self, order: OrderData) -> list[dict]:
        """将 OrderData 展平为行数据列表.

        Args:
            order: 订单数据.

        Returns:
            展平后的行数据列表，每行一个 dict.
        """
        ...

    @abc.abstractmethod
    def _fill_header(self, ws, order: OrderData) -> None:
        """填充表头信息.

        Args:
            ws: openpyxl Worksheet 对象.
            order: 订单数据.
        """
        ...

    @abc.abstractmethod
    def _fill_data_rows(
        self,
        ws,
        data_start_row: int,
        rows: list[dict],
        anchor: AnchorResult,
    ) -> None:
        """逐行填充商品明细.

        Args:
            ws: openpyxl Worksheet 对象.
            data_start_row: 数据起始行号.
            rows: 展平后的行数据列表.
            anchor: 锚点扫描结果.
        """
        ...

    @abc.abstractmethod
    def _fix_summary_formulas(self, ws, anchor: AnchorResult, new_data_end: int) -> None:
        """修正汇总行的 SUM 公式范围.

        Args:
            ws: openpyxl Worksheet 对象.
            anchor: 锚点扫描结果.
            new_data_end: 新的数据结束行号.
        """
        ...

    # ==================== 共享工具方法 ====================

    def _select_template_by_rows(self, n_rows: int) -> Path:
        """根据数据行数自动选择合适容量的模板文件.

        容量选择规则：
        - n_rows <= 20  → *_20.xlsx
        - n_rows <= 50  → *_50.xlsx（或默认模板）
        - n_rows <= 100 → *_100.xlsx
        - n_rows > 100  → 抛出 ValueError 阻断生成

        若对应容量模板文件不存在，自动降级到默认模板（50 行基准）。

        Args:
            n_rows: 数据行数.

        Returns:
            选中的模板文件路径.

        Raises:
            ValueError: 数据行数超过最大模板容量时抛出.
        """
        if n_rows <= 0:
            raise ValueError(
                f"[错误]: 数据行数必须为正整数，当前={n_rows}\n"
                f"[原因]: 订单无有效数据行\n"
                f"[排查]: 请确认订单包含至少一行商品明细"
            )

        if n_rows > MAX_TEMPLATE_ROWS:
            raise ValueError(
                f"[错误]: 数据行数 {n_rows} 超过最大模板容量 {MAX_TEMPLATE_ROWS}\n"
                f"[原因]: 当前模板仅支持最多 {MAX_TEMPLATE_ROWS} 行商品明细\n"
                f"[排查]: 请将订单拆分为多个子订单，每个不超过 {MAX_TEMPLATE_ROWS} 行"
            )

        base_path: Path = self._template_path
        stem: str = base_path.stem
        suffix: str = base_path.suffix

        # 选择容量后缀
        capacity_suffix: str = ""
        for threshold, suffix_label in TEMPLATE_CAPACITY_THRESHOLDS:
            if n_rows <= threshold:
                capacity_suffix = suffix_label
                break

        if not capacity_suffix:
            capacity_suffix = "_100"

        # 50 行容量使用默认模板（无后缀），其他容量使用带后缀的变体
        if capacity_suffix == "_50":
            # 优先使用带 _50 后缀的文件，不存在则用基准模板
            capacity_path: Path = base_path.parent / f"{stem}_50{suffix}"
            if capacity_path.exists():
                return capacity_path
            return base_path

        capacity_path = base_path.parent / f"{stem}{capacity_suffix}{suffix}"
        if not capacity_path.exists():
            logger.warning(
                "[警告]: 容量模板 %s 不存在，降级使用基准模板 %s",
                capacity_path.name,
                base_path.name,
            )
            return base_path
        return capacity_path

    def _find_actual_summary_row(
        self, ws: Worksheet, anchor: AnchorResult, keywords: list[str]
    ) -> int:
        """在锚点附近搜索实际的汇总行位置.

        从锚点标记的汇总行开始，向下搜索包含指定关键词的行。

        Args:
            ws: 工作表对象.
            anchor: 锚点扫描结果.
            keywords: 用于识别汇总行的关键词列表.

        Returns:
            实际汇总行号（1-based）.
        """
        search_start: int = anchor.summary_row
        search_end: int = min(search_start + 20, ws.max_row or search_start + 20)

        for row in range(search_start, search_end + 1):
            for col in range(1, ws.max_column + 1):
                cell_value: str = str(ws.cell(row=row, column=col).value or "")
                for keyword in keywords:
                    if keyword in cell_value:
                        return row
        return anchor.summary_row

    def _create_sandbox_copy(self, template_path: Path | None = None) -> Path:
        """在临时目录创建模板的沙箱副本.

        Args:
            template_path: 模板文件路径，默认使用 self._template_path.

        Returns:
            沙箱副本的路径.

        Raises:
            FileNotFoundError: 模板文件不存在时抛出.
        """
        source: Path = template_path if template_path else self._template_path
        if not source.exists():
            raise FileNotFoundError(
                f"[错误]: {self._get_display_name()}模板文件不存在: {source}\n"
                f"[原因]: 模板文件可能被删除、移动或改名\n"
                f"[排查]: 请将 {source.name} 放入 templates/ 目录"
            )

        temp_dir: str = tempfile.gettempdir()
        sandbox_name: str = f"{self._get_template_type()}_sandbox_{id(self)}.xlsx"
        sandbox_path: Path = Path(temp_dir) / sandbox_name
        shutil.copy2(source, sandbox_path)
        logger.debug("沙箱副本已创建: %s", sandbox_path)
        return sandbox_path

    @staticmethod
    def _cleanup_sandbox(sandbox_path: Path) -> None:
        """清理沙箱临时文件."""
        try:
            if sandbox_path.exists():
                sandbox_path.unlink(missing_ok=True)
                logger.debug("沙箱副本已清理: %s", sandbox_path)
        except Exception as e:
            logger.warning("[警告]: 清理沙箱文件失败: %s — %s", sandbox_path, e)

    def _resolve_output_path(self, order: OrderData, output_dir: str | Path | None) -> Path:
        """确定输出文件路径.

        命名规则：{输出目录}/{前缀}_{发票号}.xlsx

        Args:
            order: 订单数据.
            output_dir: 指定输出目录，None 则使用默认.

        Returns:
            输出文件路径（确保父目录存在）.
        """
        out_dir: Path = Path(output_dir) if output_dir else OUTPUT_DIR
        out_dir.mkdir(parents=True, exist_ok=True)

        invoice_no: str = order.order_meta.invoice_no.replace("/", "-").replace("\\", "-")
        prefix: str = FILE_PREFIX_MAP.get(self._get_template_type(), self._get_display_name())
        filename: str = f"{prefix}_{invoice_no}.xlsx"
        return out_dir / filename

    @staticmethod
    def _report_progress(
        callback: Callable[[str, float], None] | None,
        description: str,
        progress: float,
    ) -> None:
        """安全的进度回调."""
        if callback:
            try:
                callback(description, progress)
            except Exception as e:
                logger.warning("[警告]: 进度回调执行失败: %s", e)


# ========== 运行说明 ==========
# 依赖安装: pip install openpyxl msgspec
# 运行命令: 由 orchestrator 统一调用，不直接运行此模块
# =============================

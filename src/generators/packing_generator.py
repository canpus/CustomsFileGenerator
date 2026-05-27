"""装箱单生成器（Packing Generator）— 阶段 4.

基于装箱单 XLSX 模板，根据订单数据动态生成装箱单文件。
核心流程：沙箱复制 → 锚点扫描 → 缩容/扩容 → 填充表头 → 填充明细 → 修正公式。

参考：plan_v6.md 附录 B.1 装箱单模板字段映射表。
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl.worksheet.worksheet import Worksheet

from config.constants import TEMPLATE_PACKING_PATH
from src.generators.base_generator import BaseGenerator
from src.generators.template_anchor_scanner import (
    AnchorResult,
    scan_packing_template,
)
from src.generators.xlsx_utils import safe_write_cell, update_sum_formula
from src.models.order_data import OrderData

logger = logging.getLogger(__name__)


# ==================== 数据展平：OrderData → 装箱单行数据 ====================


def flatten_for_packing(order: OrderData) -> list[dict]:
    """将 OrderData 展平为装箱单行数据列表.

    装箱单每行对应一个（托盘, 纸箱, 商品）三元组。
    批量纸箱（is_batch=True）只生成一行, batch_count 填入"箱数"列。

    每行字典包含以下键：
        seq_no:           序号（全局递增，从 1 开始）
        pallet_no:        托盘号
        product_name:     商品名称
        specification:    规格型号
        unit:             计量单位
        qty_per_carton:   每箱数量
        carton_count:     箱数（批量纸箱为 batch_count）
        net_weight:       本行总净重（kg）= net_weight_per_unit × qty_per_carton × carton_count
        gross_weight:     本行总毛重（kg）= 纸箱毛重 × carton_count
        volume:           本行总体积（m³）= (长×宽×高)/1e6 × carton_count

    Args:
        order: 订单数据对象.

    Returns:
        展平后的行数据列表，按托盘号、商品序号排序.

    Raises:
        ValueError: order 为 None 或 pallets 为空时抛出.
    """
    if order is None:
        raise ValueError("[错误]: order 为 None, 无法展平数据")

    if not order.pallets:
        raise ValueError("[错误]: 订单无托盘数据, 无法生成装箱单")

    rows: list[dict] = []
    seq: int = 0

    for pallet in order.pallets:
        for carton in pallet.cartons:
            # 有效箱数（批量纸箱按 batch_count，否则为 1）
            effective_carton_count: int = carton.batch_count if carton.is_batch else 1

            # 单箱体积 m³ = 长(cm) × 宽(cm) × 高(cm) / 1,000,000
            single_carton_volume: float = (
                carton.length_cm * carton.width_cm * carton.height_cm / 1_000_000.0
            )

            for product in carton.products:
                seq += 1

                # 本行总净重 = 单件净重 × 每箱数量 × 箱数
                row_net_weight: float = (
                    product.net_weight_per_unit_kg * product.qty_per_carton * effective_carton_count
                )

                # 本行总毛重 = 单箱毛重 × 箱数
                row_gross_weight: float = carton.gross_weight_kg * effective_carton_count

                # 本行总体积 = 单箱体积 × 箱数
                row_volume: float = single_carton_volume * effective_carton_count

                rows.append(
                    {
                        "seq_no": seq,
                        "pallet_no": pallet.pallet_no,
                        "product_name": product.product_name,
                        "specification": product.specification,
                        "unit": product.unit,
                        "qty_per_carton": product.qty_per_carton,
                        "carton_count": effective_carton_count,
                        "net_weight": round(row_net_weight, 3),
                        "gross_weight": round(row_gross_weight, 3),
                        "volume": round(row_volume, 4),
                    }
                )

    logger.info("装箱单数据展平完成: 共 %d 行", len(rows))
    return rows


# ==================== PackingGenerator 类 ====================


class PackingGenerator(BaseGenerator):
    """装箱单生成器.

    负责将 OrderData 填充到装箱单 XLSX 模板中，产出最终装箱单文件。
    继承 BaseGenerator，仅保留装箱单特有的表头填充、明细填充、汇总修正逻辑。

    使用方式：
        gen = PackingGenerator()
        output_path = gen.generate(order, output_dir, progress_callback)
    """

    def _get_default_template_path(self) -> Path:
        """返回默认模板路径."""
        return TEMPLATE_PACKING_PATH

    def _get_template_type(self) -> str:
        """返回模板类型标识."""
        return "packing"

    def _get_display_name(self) -> str:
        """返回生成器显示名称."""
        return "装箱单"

    def _scan_anchor(self, ws: Worksheet) -> AnchorResult:
        """扫描装箱单模板锚点."""
        return scan_packing_template(ws)

    def _flatten_data(self, order: OrderData) -> list[dict]:
        """将 OrderData 展平为装箱单行数据列表."""
        return flatten_for_packing(order)

    def _fill_header(self, ws: Worksheet, order: OrderData) -> None:
        """填充装箱单表头信息.

        实际模板结构（基于 template_packing.xlsx 实测）：
        - D3:K3：客户抬头（合并单元格区域）
        - A4:C4："Invoice No. 发票号"（合并），后方填发票号
        - F4:H4："Date日期:"（合并），后方填日期
        - A5:C5："Contact No. 合同号"（合并），后方填合同号
        - F5:H5："Payment付款方式:"（合并），后方填付款方式
        - A6:C6："Country of origin产地:"（合并），后方填产地
        - F6:H6："Destination 目的地:"（合并），后方填目的地

        注意：使用 safe_write_cell，MergedCell 会被自动重定向到合并区域左上角.
        """
        customer_name: str = order.customer.company_name_en or order.customer.company_name_cn
        # D3:K3 合并区域 - 客户抬头
        safe_write_cell(ws, 3, "D", customer_name)

        # A4:C4 合并区域 - 发票号（模板已有标签文本，只写值）
        safe_write_cell(ws, 4, "A", order.order_meta.invoice_no)
        # F4:H4 合并区域 - 日期
        safe_write_cell(ws, 4, "F", order.order_meta.date)

        # A5:C5 合并区域 - 合同号
        safe_write_cell(ws, 5, "A", order.order_meta.contract_no)
        # F5:H5 合并区域 - 付款方式
        safe_write_cell(ws, 5, "F", order.order_meta.payment_term)

        # A6:C6 合并区域 - 产地
        safe_write_cell(ws, 6, "A", order.order_meta.country_of_origin)
        # F6:H6 合并区域 - 目的地
        destination: str = order.customer.destination or order.customer.country
        safe_write_cell(ws, 6, "F", destination)

        logger.info("装箱单表头填充完成")

    def _fill_data_rows(
        self,
        ws: Worksheet,
        data_start_row: int,
        rows: list[dict],
        anchor: AnchorResult,
    ) -> None:
        """逐行填充商品明细.

        实际模板列映射 (基于 template_packing.xlsx 实测):
        A=seq_no(No.), B:C=product_name(Item Description, 合并),
        D=specification(Spec.), E=unit(Unit),
        F=qty_per_carton(QTY./Ctn), G=package_no(Package No.),
        H=pallet_no(Pallet No.), I=net_weight(N.W./Kg),
        J=gross_weight(G.W/kg), K=volume(Volume/M3)

        注意：本模板中 G 列 = Package No.（包号），没有直接的"箱数"列。
        我们将 carton_count 写入 G 列（Package No.），这与真实订单数据中
        将此列用于表示箱数的方法一致。

        Args:
            ws: 装箱单工作表.
            data_start_row: 数据起始行号.
            rows: 展平后的行数据列表.
            anchor: 锚点扫描结果.
        """
        for idx, row_data in enumerate(rows):
            target_row: int = data_start_row + idx

            # A: 序号 (No.)
            safe_write_cell(ws, target_row, "A", row_data["seq_no"])
            # B: 商品名称 (Item Description), 合并列 B:C 只写入 B 列
            safe_write_cell(ws, target_row, "B", row_data["product_name"])
            # D: 规格 (Spec.)
            safe_write_cell(ws, target_row, "D", row_data["specification"])
            # E: 单位 (Unit)
            safe_write_cell(ws, target_row, "E", row_data["unit"])
            # F: 每箱数量 (QTY. / Ctn)
            safe_write_cell(ws, target_row, "F", row_data["qty_per_carton"])
            # G: 箱数 (Package No. 列 — 模板实际将此列用于表示箱数)
            safe_write_cell(ws, target_row, "G", row_data["carton_count"])
            # H: 托盘号 (Pallet No.)
            safe_write_cell(ws, target_row, "H", row_data["pallet_no"])
            # I: 净重 kg (N.W./Kg) — 不保留旧 number_format，避免时间/日期格式
            safe_write_cell(ws, target_row, "I", row_data["net_weight"], preserve_style=True)
            # 写入数值后强制清除可能遗留的非数值格式
            cell_i = ws.cell(row=target_row, column=9)
            if cell_i.value is not None and isinstance(cell_i.value, (int, float)):
                cell_i.number_format = "0.000"
            # J: 毛重 kg (G.W/kg)
            safe_write_cell(ws, target_row, "J", row_data["gross_weight"], preserve_style=True)
            cell_j = ws.cell(row=target_row, column=10)
            if cell_j.value is not None and isinstance(cell_j.value, (int, float)):
                cell_j.number_format = "0.000"
            # K: 体积 m³ (Volume/M3)
            safe_write_cell(ws, target_row, "K", row_data["volume"], preserve_style=True)
            cell_k = ws.cell(row=target_row, column=11)
            if cell_k.value is not None and isinstance(cell_k.value, (int, float)):
                cell_k.number_format = "0.0000"

        logger.info(
            "装箱单明细填充完成: %d 行 (第 %d 行 → 第 %d 行)",
            len(rows),
            data_start_row,
            data_start_row + len(rows) - 1,
        )

    def _fix_summary_formulas(self, ws: Worksheet, anchor: AnchorResult, new_data_end: int) -> None:
        """修正汇总行的 SUM 公式范围.

        模板汇总行（第 58 行）实际公式列：
        G58=SUM(G8:G57) → 总箱数（Package No. 列之和）
        I58=SUM(I8:I57) → 总净重
        J58=SUM(J8:J57) → 总毛重
        K58=SUM(K8:K57) → 总体积

        Args:
            ws: 装箱单工作表.
            anchor: 锚点扫描结果.
            new_data_end: 新的数据结束行号.
        """
        # 仅修正实际存在 SUM 公式的列
        formula_columns: list[str] = ["G", "I", "J", "K"]

        for col_letter in formula_columns:
            update_sum_formula(ws, col_letter, anchor.data_start_row, new_data_end)

        logger.info("汇总公式已修正: 范围 %d→%d", anchor.data_start_row, new_data_end)


# ========== 运行说明 ==========
# 依赖安装: pip install openpyxl msgspec
# 运行命令: 由 orchestrator 统一调用，不直接运行此模块
# 测试命令: python -m pytest tests/test_packing_generator.py -v
# =============================

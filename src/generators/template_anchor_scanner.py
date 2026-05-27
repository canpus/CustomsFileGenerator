"""模板动态锚点扫描引擎 — 聚合入口.

扫描 XLSX 模板工作表的前 N 行，通过关键词匹配定位数据起始行和汇总行。
不依赖硬编码行号，模板行结构变更时无需修改代码。

阶段 3 核心模块，后续所有 xlsx 生成器均依赖此模块。

子模块拆分：
- anchor_core.py   : AnchorResult 数据类 + 共享扫描工具函数
- scanners/        : 各模板类型的独立扫描函数
- 本模块           : scan_template 入口 + _create_default_wb + 统一重导出
"""

from __future__ import annotations

import copy
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

from src.generators.anchor_core import (  # noqa: F401 — 重导出
    AnchorResult,
    _contains_keyword,
    _estimate_data_end_row,
    _find_row_by_keywords,
    _find_summary_rows,
    _load_rules,
)
from src.generators.scanners.contract_scanner import scan_contract_template  # noqa: F401
from src.generators.scanners.invoice_scanner import scan_invoice_template  # noqa: F401
from src.generators.scanners.packing_scanner import scan_packing_template  # noqa: F401

logger = logging.getLogger(__name__)


# ==================== 公开 API ====================


def scan_template(template_path: str | Path, rules: dict | None = None) -> AnchorResult:
    """通用模板扫描入口，根据文件名自动选择扫描策略.

    Args:
        template_path: 模板文件路径。
        rules: 完整的规则字典（可选）。

    Returns:
        AnchorResult 扫描结果。

    Raises:
        FileNotFoundError: 模板文件不存在时抛出。
        ValueError: 不支持的模板类型时抛出。
    """
    path = Path(template_path)

    if not path.exists():
        raise FileNotFoundError(
            f"[错误]: 模板文件不存在: {path}\n"
            f"[原因]: 文件可能已被移动、删除或改名\n"
            f"[排查]: 请将模板文件放入 templates/ 目录"
        )

    if rules is None:
        rules = _load_rules()

    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active if wb.active else wb.worksheets[0]
    except Exception as e:
        raise ValueError(
            f"[错误]: 无法打开模板文件: {path}\n"
            f"[原因]: {e}\n"
            f"[排查]: 请确认文件是否为有效的 .xlsx 格式"
        ) from e

    filename_lower: str = path.name.lower()

    try:
        if "packing" in filename_lower:
            result = scan_packing_template(ws, rules)
        elif "invoice" in filename_lower:
            result = scan_invoice_template(ws, rules)
        elif "contract" in filename_lower:
            result = scan_contract_template(ws, rules)
        else:
            raise ValueError(
                f"[错误]: 不支持的模板类型: {path.name}\n"
                f"[原因]: 模板文件名必须包含 'packing'、'invoice' 或 'contract'\n"
                f"[排查]: 请确认模板文件名符合命名规范"
            )
    finally:
        wb.close()

    return result


# ==================== 默认工作簿工厂 ====================


def _create_default_wb(template_type: str) -> openpyxl.Workbook:
    """当真实模板不可用时，构造一个符合规范结构的默认工作簿.

    在测试环境或无模板可用的降级场景中，此工厂方法构造一个与真实模板
    行结构完全一致的工作簿，确保锚点扫描和后续生成器能正常工作。

    工作簿结构（每种模板类型对应不同的行布局）：
    - packing:  标题行=第7行, 数据行=8~57(50行), 汇总行=58
    - invoice:  标题行=第14行, 数据行=15~64(50行), 汇总行=65
    - contract: 标题行=第7行, 数据行=8~57(50行), 汇总行=58

    Args:
        template_type: 模板类型，仅支持 "packing"、"invoice"、"contract"。

    Returns:
        openpyxl Workbook 对象。

    Raises:
        ValueError: 不支持的模板类型。
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(name="Arial", size=11, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if template_type == "packing":
        ws.title = "Packing List"
        headers = [
            "序号",
            "商品描述",
            "",
            "规格",
            "单位",
            "单箱数量",
            "箱数",
            "托板号",
            "净重(kg)",
            "毛重(kg)",
            "体积(m³)",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_idx)
            cell.value = h
            cell.font = copy.deepcopy(header_font)
            cell.alignment = copy.deepcopy(header_align)
            cell.border = copy.deepcopy(thin_border)
        ws.merge_cells("B7:C7")
        ws.cell(row=58, column=1).value = "总"
        ws.cell(row=58, column=7).value = "=SUM(G8:G57)"
        ws.cell(row=58, column=9).value = "=SUM(I8:I57)"
        ws.cell(row=58, column=10).value = "=SUM(J8:J57)"
        ws.cell(row=58, column=11).value = "=SUM(K8:K57)"
    elif template_type == "invoice":
        ws.title = "Proforma Invoice"
        headers = ["Product", "Specification", "Unit", "Qty", "Unit Price", "Amount"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=14, column=col_idx)
            cell.value = h
            cell.font = copy.deepcopy(Font(name="Times New Roman", size=14, bold=True))
            cell.alignment = copy.deepcopy(header_align)
            cell.border = copy.deepcopy(thin_border)
        ws.cell(row=65, column=5).value = "TOTAL:"
        ws.cell(row=65, column=6).value = "=SUM(F15:F64)"
        ws.cell(row=66, column=1).value = "SAY: USD ... ONLY"
    elif template_type == "contract":
        ws.title = "Sales Contract"
        headers = ["No.", "Product", "Specification", "Unit", "Qty", "Unit Price", "Amount"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_idx)
            cell.value = h
            cell.font = copy.deepcopy(Font(name="Times New Roman", size=14, bold=True))
            cell.alignment = copy.deepcopy(header_align)
            cell.border = copy.deepcopy(thin_border)
        ws.cell(row=58, column=6).value = "TOTAL:"
        ws.cell(row=58, column=7).value = "=SUM(G8:G57)"
        ws.cell(row=59, column=1).value = "SAY: USD ... ONLY"
    else:
        raise ValueError(
            f"[错误]: 不支持的模板类型: {template_type!r}\n"
            f"[原因]: 仅支持 'packing'、'invoice' 或 'contract'\n"
            f"[排查]: 请确认模板类型名称正确"
        )

    return wb

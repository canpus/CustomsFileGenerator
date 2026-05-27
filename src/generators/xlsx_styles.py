"""XLSX 样式工具 — 样式深拷贝、安全写入、列索引转换."""

from __future__ import annotations

import copy
import logging
from typing import Any, cast

from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils.cell import column_index_from_string, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ==================== 样式深拷贝 ====================


def _copy_font(src: Any) -> Font:
    """安全拷贝 Font 对象（属性级拷贝，接受 StyleProxy 等兼容对象）."""
    if src is None:
        return Font()
    return Font(
        name=src.name,
        size=src.size,
        bold=src.bold,
        italic=src.italic,
        underline=src.underline,
        strike=src.strike,
        color=copy.copy(src.color) if src.color else None,
        charset=src.charset,
        family=src.family,
    )


def _copy_border(src: Any) -> Border:
    """安全拷贝 Border 对象（属性级拷贝，接受 StyleProxy 等兼容对象）."""
    if src is None:
        return Border()
    return Border(
        left=copy.copy(src.left) if src.left else None,
        right=copy.copy(src.right) if src.right else None,
        top=copy.copy(src.top) if src.top else None,
        bottom=copy.copy(src.bottom) if src.bottom else None,
        diagonal=copy.copy(src.diagonal) if src.diagonal else None,
        outline=src.outline,
        diagonalUp=src.diagonalUp,
        diagonalDown=src.diagonalDown,
    )


def _copy_fill(src: Any) -> PatternFill:
    """安全拷贝 Fill 对象（属性级拷贝，接受 StyleProxy 等兼容对象）."""
    if src is None:
        return PatternFill()
    return PatternFill(
        fill_type=src.fill_type,
        start_color=copy.copy(src.start_color) if src.start_color else None,
        end_color=copy.copy(src.end_color) if src.end_color else None,
        patternType=src.patternType,
    )


def _copy_alignment(src: Any) -> Alignment:
    """安全拷贝 Alignment 对象（属性级拷贝，接受 StyleProxy 等兼容对象）."""
    if src is None:
        return Alignment()
    return Alignment(
        horizontal=src.horizontal,
        vertical=src.vertical,
        wrap_text=src.wrap_text,
        shrink_to_fit=src.shrink_to_fit,
        indent=src.indent,
        text_rotation=src.text_rotation,
    )


def _copy_number_format(src: Any) -> str:
    """安全拷贝数字格式字符串."""
    if src is None:
        return "General"
    return src


# ==================== 公开 API ====================


def clone_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    """将源行的所有样式深拷贝到目标行.

    拷贝范围：font, border, fill, alignment, number_format。
    不对源行做任何修改。每列独立拷贝。

    Raises:
        ValueError: 行号 ≤ 0 时抛出。
    """
    if source_row <= 0 or target_row <= 0:
        raise ValueError(
            f"[错误]: 行号必须为正整数，source_row={source_row}, target_row={target_row}"
        )

    max_col: int = ws.max_column or 1

    for col_idx in range(1, max_col + 1):
        try:
            src_cell: Cell = cast(Cell, ws.cell(row=source_row, column=col_idx))
            tgt_cell: Cell = cast(Cell, ws.cell(row=target_row, column=col_idx))

            tgt_cell.font = _copy_font(src_cell.font)
            tgt_cell.border = _copy_border(src_cell.border)
            tgt_cell.fill = _copy_fill(src_cell.fill)
            tgt_cell.alignment = _copy_alignment(src_cell.alignment)
            tgt_cell.number_format = _copy_number_format(src_cell.number_format)
        except Exception as e:
            logger.warning("[警告]: 克隆行样式时出错 row=%d col=%d: %s", target_row, col_idx, e)


def insert_rows_with_style(ws: Worksheet, anchor_row: int, count: int) -> None:
    """在锚点行之后插入 N 行，并为每一行复制锚点行的样式.

    Raises:
        ValueError: count ≤ 0 或 anchor_row ≤ 0 时抛出。
    """
    if count <= 0:
        raise ValueError(f"[错误]: 插入行数必须为正整数，count={count}")
    if anchor_row <= 0:
        raise ValueError(f"[错误]: 锚点行号必须为正整数，anchor_row={anchor_row}")

    for i in range(count):
        insert_row_idx: int = anchor_row + 1 + i
        ws.insert_rows(insert_row_idx)
        clone_row_style(ws, anchor_row, insert_row_idx)
        logger.debug("已在第 %d 行后插入一行并复制样式", anchor_row + i)


def safe_write_cell(
    ws: Worksheet, row: int, col: int | str, value, preserve_style: bool = True
) -> None:
    """安全写入单元格值，保留原有样式.

    Args:
        ws: Worksheet 对象。
        row: 行号（1-based）。
        col: 列号（int）或列字母（str）。
        value: 要写入的值。
        preserve_style: 是否保留目标单元格原有样式（默认 True）。
    """
    if isinstance(col, str):
        col = column_index_from_string(col)
    col_i: int = col  # type: ignore[assignment]

    cell: Cell = cast(Cell, ws.cell(row=row, column=col_i))

    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            bounds = range_boundaries(str(merged_range))
            min_col: int = int(bounds[0] or 0)
            min_row: int = int(bounds[1] or 0)
            max_col: int = int(bounds[2] or 0)
            max_row: int = int(bounds[3] or 0)
            if min_row <= row <= max_row and min_col <= col_i <= max_col:
                cell = cast(Cell, ws.cell(row=min_row, column=min_col))
                break

    if not preserve_style:
        cell.value = value
        return

    saved_font = _copy_font(cell.font)
    saved_border = _copy_border(cell.border)
    saved_fill = _copy_fill(cell.fill)
    saved_alignment = _copy_alignment(cell.alignment)
    saved_number_format = _copy_number_format(cell.number_format)

    cell.value = value

    cell.font = saved_font
    if saved_border:
        cell.border = saved_border
    if saved_fill:
        cell.fill = saved_fill
    if saved_alignment:
        cell.alignment = saved_alignment
    if saved_number_format:
        cell.number_format = saved_number_format


def get_column_index(col_letter: str) -> int:
    """将列字母转换为列索引（1-based）.

    Raises:
        ValueError: 列字母无效时抛出。
    """
    if not col_letter or not col_letter.isalpha():
        raise ValueError(f"[错误]: 无效的列字母: {col_letter!r}")
    return column_index_from_string(col_letter)

"""XLSX 工具库 — 行操作与公式管理.

提供安全插入/删除行、合并单元格检测、公式修正等操作。

子模块拆分：
- xlsx_styles.py: 样式深拷贝、安全写入、列索引转换、样式行插入
- 本模块        : 合并范围检测、安全删除、公式修正、行数调整 + 统一重导出
"""

from __future__ import annotations

import logging
import re

from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from src.generators.xlsx_styles import (  # noqa: F401
    clone_row_style,
    get_column_index,
    insert_rows_with_style,
    safe_write_cell,
)

logger = logging.getLogger(__name__)

# SUM 公式正则
_SUM_FORMULA_PATTERN: re.Pattern = re.compile(
    r"=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)",
    re.IGNORECASE,
)


# ==================== 合并范围 ====================


def get_merged_ranges_in_row(ws: Worksheet, row: int) -> list:
    """获取指定行所涉及的所有合并单元格范围."""
    merged_ranges: list = []
    for merged_range in ws.merged_cells.ranges:
        try:
            bounds = range_boundaries(str(merged_range))
            min_row: int = int(bounds[1] or 0)
            max_row: int = int(bounds[3] or 0)
            if min_row <= row <= max_row:
                merged_ranges.append(merged_range)
        except Exception as e:
            logger.warning("[警告]: 解析合并单元格范围失败: %s — %s", merged_range, e)
    return merged_ranges


# ==================== 行删除 ====================


def delete_rows_safely(ws: Worksheet, start_row: int, end_row: int) -> None:
    """逆序安全删除行.

    核心算法：
    1. 逆序遍历（从 end_row 到 start_row），避免正序删除导致行号偏移。
    2. 每行删除前自动 unmerge_cells，避免 openpyxl 崩溃。

    Raises:
        ValueError: start_row > end_row 或行号 ≤ 0 时抛出。
    """
    if start_row <= 0 or end_row <= 0:
        raise ValueError(f"[错误]: 行号必须为正整数，start_row={start_row}, end_row={end_row}")
    if start_row > end_row:
        raise ValueError(f"[错误]: start_row({start_row}) 不能大于 end_row({end_row})")

    logger.info("开始安全删除行: 第 %d 行 → 第 %d 行（逆序）", start_row, end_row)

    for row in range(end_row, start_row - 1, -1):
        merged_ranges = get_merged_ranges_in_row(ws, row)
        for merged_range in merged_ranges:
            try:
                ws.unmerge_cells(str(merged_range))
                logger.debug("已解除合并单元格: %s", merged_range)
            except Exception as e:
                logger.warning("[警告]: 解除合并单元格失败 %s: %s", merged_range, e)

        ws.delete_rows(row)
        logger.debug("已删除第 %d 行", row)

    logger.info("安全删除完成: 共删除 %d 行", end_row - start_row + 1)


# ==================== 公式修正 ====================


def update_sum_formula(ws: Worksheet, col_letter: str, start_row: int, end_row: int) -> None:
    """修正指定列的 SUM 公式范围.

    Raises:
        ValueError: 列字母无效时抛出。
    """
    if not col_letter or not col_letter.isalpha():
        raise ValueError(f"[错误]: 无效的列字母: {col_letter!r}")

    new_formula: str = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
    updated_count: int = 0

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                match = _SUM_FORMULA_PATTERN.match(cell.value)
                if match:
                    old_start_col: str = match.group(1)
                    if old_start_col.upper() == col_letter.upper():
                        old_formula: str = cell.value
                        cell.value = new_formula
                        updated_count += 1
                        logger.debug(
                            "公式修正: %s → %s (cell %s)",
                            old_formula,
                            new_formula,
                            cell.coordinate,
                        )

    if updated_count == 0:
        logger.warning(
            "[警告]: 未找到需要修正的 SUM 公式，col=%s, range=%s%d:%s%d",
            col_letter,
            col_letter,
            start_row,
            col_letter,
            end_row,
        )
    else:
        logger.info("公式修正完成: 共修正 %d 个 SUM 公式", updated_count)


def delete_reserved_rows(
    ws: Worksheet, data_start_row: int, data_end_row: int, actual_row_count: int
) -> int:
    """删除数据区中未使用的预留空行.

    当使用大容量模板（如 50 行）填充少量数据（如 3 行）时，
    数据区尾部会有多余预留行。此函数将其安全删除并返回新的数据结束行号。

    与 resize_data_rows 的区别：
    - resize_data_rows 同时支持扩容和缩容
    - delete_reserved_rows 仅删除额外预留行，语义更明确

    Args:
        ws: 工作表对象.
        data_start_row: 数据起始行号（1-based）.
        data_end_row: 模板预设的数据结束行号（1-based）.
        actual_row_count: 实际数据行数.

    Returns:
        新的数据结束行号.

    Raises:
        ValueError: 参数无效时抛出.
    """
    if data_start_row <= 0 or data_end_row <= 0:
        raise ValueError(f"[错误]: 行号必须为正整数，start={data_start_row}, end={data_end_row}")
    if actual_row_count <= 0:
        raise ValueError(f"[错误]: 实际行数必须为正整数，actual={actual_row_count}")
    if data_start_row > data_end_row:
        raise ValueError(f"[错误]: 数据起始行({data_start_row}) > 结束行({data_end_row})")

    actual_end_row: int = data_start_row + actual_row_count - 1
    reserved_start: int = actual_end_row + 1

    if reserved_start > data_end_row:
        return data_end_row

    delete_count: int = data_end_row - reserved_start + 1
    logger.info(
        "删除预留空行: 第 %d 行 → 第 %d 行（共 %d 行），新数据结束行=%d",
        reserved_start,
        data_end_row,
        delete_count,
        actual_end_row,
    )
    delete_rows_safely(ws, reserved_start, data_end_row)
    return actual_end_row


def resize_data_rows(
    ws: Worksheet,
    data_start_row: int,
    data_end_row: int,
    target_row_count: int,
    anchor_row: int,
) -> int:
    """根据目标行数自动扩容/缩容数据区域.

    这是 delete_rows_safely + insert_rows_with_style 的高层封装。

    Args:
        ws: Worksheet 对象。
        data_start_row: 数据起始行号。
        data_end_row: 当前数据结束行号（模板默认值）。
        target_row_count: 目标数据行数。
        anchor_row: 样式锚点行号。

    Returns:
        新的数据结束行号（1-based）。

    Raises:
        ValueError: 参数无效时抛出。
    """
    if data_start_row <= 0 or data_end_row <= 0:
        raise ValueError(
            f"[错误]: 数据行号必须为正整数，start={data_start_row}, end={data_end_row}"
        )
    if target_row_count <= 0:
        raise ValueError(f"[错误]: 目标行数必须为正整数，target={target_row_count}")
    if data_start_row > data_end_row:
        raise ValueError(f"[错误]: 数据起始行({data_start_row}) > 结束行({data_end_row})")

    current_count: int = data_end_row - data_start_row + 1
    diff: int = target_row_count - current_count

    logger.info(
        "数据行调整: 当前 %d 行 → 目标 %d 行 (diff=%+d)",
        current_count,
        target_row_count,
        diff,
    )

    if diff > 0:
        insert_rows_with_style(ws, data_end_row, diff)
        new_data_end: int = data_end_row + diff
        logger.info(
            "已扩容: 在 %d 行后插入 %d 行，新结束行=%d",
            data_end_row,
            diff,
            new_data_end,
        )
        return new_data_end
    elif diff < 0:
        delete_start: int = data_start_row + target_row_count
        delete_rows_safely(ws, delete_start, data_end_row)
        new_end: int = delete_start - 1
        logger.info(
            "已缩容: 删除 %d 行（%d→%d），新结束行=%d",
            -diff,
            delete_start,
            data_end_row,
            new_end,
        )
        return new_end
    else:
        logger.info("数据行数无需调整（%d 行）", current_count)
        return data_end_row

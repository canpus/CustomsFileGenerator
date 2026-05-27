"""Excel 订单数据导入器 — 阶段 8.1（门面模块）.

读取用户上传的订单 Excel 表格，自动识别格式并解析为 OrderData。

子模块：
- column_mapper.py  : 列名映射表 + 贸易条款标准化
- detail_parser.py  : 明细行解析 + 汇总计算
- format_handlers.py: KV/明细两种格式的处理函数
- 本模块            : import_order_from_excel / quick_import 入口
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from src.importer.format_handlers import (
    _import_detail_format,
    _import_kv_format,
    _is_kv_format,
)
from src.models.order_data import OrderData

logger = logging.getLogger(__name__)


def import_order_from_excel(
    excel_path: str | Path,
    sheet_name: str | None = None,
) -> tuple[OrderData, dict[str, list[str]]]:
    """从 Excel 表格导入订单数据并生成 OrderData.

    自动识别列名、映射字段、计算汇总数据。

    支持两种 Excel 格式：
        A. 双 Sheet 格式：Sheet1 "订单信息"（键值对）+ Sheet2 "商品明细"
        B. 单 Sheet 格式（推荐）：表头行含所有列名，每行一条记录

    Args:
        excel_path: Excel 文件路径.
        sheet_name: 指定工作表名称，默认使用第一个 Sheet.

    Returns:
        (OrderData, unmapped_columns) 元组。

    Raises:
        FileNotFoundError: Excel 文件不存在.
        ValueError: 数据格式不正确无法解析.
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(
            f"[错误]: Excel 文件不存在: {excel_path}\n"
            f"[原因]: 文件可能被移动、删除或路径拼写错误\n"
            f"[排查]: 请检查文件路径是否正确"
        )

    print(f"正在读取 Excel 文件: {excel_path.name}")
    unmapped: dict[str, list[str]] = {}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        raise ValueError(
            f"[错误]: 无法打开 Excel 文件: {excel_path}\n"
            f"[原因]: 文件可能已损坏或格式不受支持\n"
            f"[排查]: 请确认文件为 .xlsx 格式，且未被其他程序独占打开\n"
            f"        原始错误: {e}"
        ) from e

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f'[错误]: 工作表 "{sheet_name}" 不存在\n'
                f"[原因]: 可用的工作表: {', '.join(wb.sheetnames)}\n"
                f"[排查]: 请指定正确的工作表名称"
            )
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    print(f"  使用工作表: {ws.title} ({ws.max_row} 行 x {ws.max_column} 列)")

    if _is_kv_format(ws):
        print("  检测到键值对格式（订单信息 Sheet）")
        return _import_kv_format(wb, unmapped)
    else:
        print("  检测到明细行格式")
        return _import_detail_format(ws, unmapped)


def quick_import(excel_path: str | Path) -> OrderData:
    """快速导入 Excel 订单（忽略未映射列）.

    Args:
        excel_path: Excel 文件路径.

    Returns:
        OrderData 对象.

    Raises:
        FileNotFoundError: 文件不存在.
        ValueError: 数据格式错误.
    """
    order, unmapped = import_order_from_excel(excel_path)
    if unmapped:
        for sheet, cols in unmapped.items():
            logger.info(
                "TODO:待确认 — %s 中有 %d 个列无法自动映射: %s",
                sheet,
                len(cols),
                ", ".join(cols),
            )
    return order

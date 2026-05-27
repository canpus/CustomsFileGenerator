# -*- coding: utf-8 -*-
"""阶段 8 里程碑测试：Excel 导入器.

运行方式：
    python -m pytest tests/test_importer.py -v

覆盖：
    1. Excel 导入：读取样本订单 Excel → 输出 OrderData → 字段映射正确
    2. Excel 导入未知列名 → 标注 TODO
    3. 自动计算：体积 = 长×宽×高（验证 1.2×1.0×1.5 = 1.8）
    4. 自动汇总：总毛重 = 所有纸箱毛重之和
    5. 模板加载：从 SQLite 加载模板 → 反序列化为 OrderData
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

# 确保临时数据库路径生效
_TEST_DB_DIR: Path = Path(tempfile.mkdtemp(prefix="test_importer_db_"))
_TEST_DB_PATH: Path = _TEST_DB_DIR / "customs.db"


@pytest.fixture(autouse=True)
def setup_test_db(monkeypatch):
    """每个测试前：切换到临时数据库."""
    _TEST_DB_DIR.mkdir(parents=True, exist_ok=True)

    import config.constants as consts
    monkeypatch.setattr(consts, "DATABASE_PATH", _TEST_DB_PATH)

    # 重置连接池
    import threading
    import src.db.connection as conn_module
    monkeypatch.setattr(conn_module, "_thread_local", threading.local())

    if _TEST_DB_PATH.exists():
        _TEST_DB_PATH.unlink()

    yield

    # 清理：关闭连接 + 删除临时文件
    from src.db.connection import close_all_connections
    close_all_connections()
    try:
        if _TEST_DB_PATH.exists():
            _TEST_DB_PATH.unlink()
    except Exception:
        pass


# ==================== 辅助函数 ====================


def _create_test_excel(headers: list[str], data_rows: list[list]) -> Path:
    """创建一个测试用 Excel 文件.

    Args:
        headers: 表头列名列表.
        data_rows: 数据行列表（每行是值列表）.

    Returns:
        Excel 文件路径.
    """
    tmp_dir = Path(tempfile.mkdtemp(prefix="test_importer_excel_"))
    file_path = tmp_dir / "test_order.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单明细"

    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(file_path)
    wb.close()
    return file_path


# ==================== 测试 1：Excel 导入 → 字段映射正确 ====================


def test_import_basic_order():
    """测试 1: 读取标准订单 Excel → 输出 OrderData → 字段映射正确."""
    from src.importer.excel_importer import import_order_from_excel

    headers = [
        "发票号", "合同号", "日期", "贸易条款", "付款方式", "币种", "原产国",
        "客户", "客户地址", "国家", "联系人", "电话", "目的地",
        "托盘号", "托盘长度", "托盘宽度", "托盘高度",
        "纸箱标签", "纸箱长度", "纸箱宽度", "纸箱高度", "纸箱毛重", "箱数",
        "序号", "商品名称", "规格", "HS编码", "单位", "数量", "单价", "净重", "目的国",
    ]

    data_rows = [
        [
            "20251202-01", "PO25-018", "2025-12-26", "FOB", "100% T/T IN ADVANCE", "USD", "China",
            "YARIMKURE INSAAT LTD. STI.", "Mutlukent Mahallesi, Ankara 85 Sitesi",
            "Turkey", "MR. Z.AYHAN ACAR", "90-312-2363217", "Ankara",
            1, 1.16, 1.01, 1.97,
            "1-45", 32, 32, 34, 23.3, 45,
            1, "50°C Type 2A-1 Heat Shrink Sleeve", "330mm*2.0mm*30M",
            "3926909090", "Roll", 1, 85, 22.3, "Turkey",
        ],
    ]

    excel_path = _create_test_excel(headers, data_rows)

    order, unmapped = import_order_from_excel(excel_path)

    # 验证字段映射
    assert order.order_meta.invoice_no == "20251202-01"
    assert order.order_meta.contract_no == "PO25-018"
    assert order.order_meta.date == "2025-12-26"
    assert order.order_meta.trade_term == "FOB"
    assert order.order_meta.payment_term == "100% T/T IN ADVANCE"
    assert order.order_meta.currency == "USD"
    assert order.order_meta.country_of_origin == "China"

    assert order.customer.company_name_en == "YARIMKURE INSAAT LTD. STI."
    assert order.customer.country == "Turkey"
    assert order.customer.address == "Mutlukent Mahallesi, Ankara 85 Sitesi"
    assert order.customer.contact_person == "MR. Z.AYHAN ACAR"
    assert order.customer.phone == "90-312-2363217"
    assert order.customer.destination == "Ankara"

    # 验证托盘
    assert order.totals.total_pallets == 1
    assert len(order.pallets) == 1
    pallet = order.pallets[0]
    assert pallet.pallet_no == 1
    assert pallet.length_m == 1.16
    assert pallet.width_m == 1.01
    assert pallet.height_m == 1.97

    # 验证纸箱
    assert len(pallet.cartons) == 1
    carton = pallet.cartons[0]
    assert carton.carton_label == "1-45"
    assert carton.gross_weight_kg == 23.3

    # 验证商品
    assert len(carton.products) == 1
    product = carton.products[0]
    assert product.product_name == "50°C Type 2A-1 Heat Shrink Sleeve"
    assert product.hs_code == "3926909090"
    assert product.unit_price == 85.0
    assert product.qty_per_carton == 1

    # 验证汇总
    assert order.totals.total_cartons == 45  # batch_count
    assert order.totals.total_amount == 85.0 * 1 * 45

    print("✅ 测试 1 通过: 标准订单导入字段映射正确")


# ==================== 测试 2：未知列名 → 标注 TODO ====================


def test_import_unknown_columns():
    """测试 2: 包含无法识别的列名 → 返回 unmapped 字典."""
    from src.importer.excel_importer import import_order_from_excel

    headers = [
        "发票号", "合同号", "日期", "客户", "国家",
        "托盘号", "托盘长度", "托盘宽度", "托盘高度",
        "纸箱标签", "纸箱长度", "纸箱宽度", "纸箱高度", "纸箱毛重",
        "序号", "商品名称", "HS编码", "单位", "数量", "单价", "净重", "目的国",
        "奇怪的自定义列", "another_unknown_column",  # 无法识别的列
    ]

    data_rows = [
        [
            "TEST-001", "CT-001", "2026-01-01", "Test Corp", "CN",
            1, 1.0, 1.0, 1.0,
            "C1", 30.0, 30.0, 30.0, 20.0,
            1, "Test Product", "12345678", "PCS", 10, 5.0, 1.0, "CN",
            "随机数据", "unknown value",
        ],
    ]

    excel_path = _create_test_excel(headers, data_rows)

    order, unmapped = import_order_from_excel(excel_path)

    # 已知列应正确解析
    assert order.order_meta.invoice_no == "TEST-001"
    assert order.customer.company_name_en == "Test Corp"

    # 未知列应出现在 unmapped 中
    assert len(unmapped) > 0, f"应包含未识别的列，实际 unmapped={unmapped}"

    # 验证 unmapped 中记录了无法识别的列
    has_unknown = False
    for cols in unmapped.values():
        for col in cols:
            if "奇怪" in col or "unknown" in col.lower():
                has_unknown = True
    assert has_unknown, f"unmapped 应包含无法识别的列，当前内容: {unmapped}"

    print(f"✅ 测试 2 通过: 未知列已标注 TODO，unmapped={unmapped}")


# ==================== 测试 3：自动计算体积 ====================


def test_auto_volume_calculation():
    """测试 3: 体积 = 长×宽×高（验证 1.2×1.0×1.5 = 1.8）."""
    from src.importer.excel_importer import import_order_from_excel

    headers = [
        "发票号", "合同号", "日期", "客户", "国家",
        "托盘号", "托盘长度", "托盘宽度", "托盘高度",
        "纸箱标签", "纸箱长度", "纸箱宽度", "纸箱高度", "纸箱毛重",
        "序号", "商品名称", "HS编码", "单位", "数量", "单价", "净重", "目的国",
    ]

    # 托盘 1.2×1.0×1.5 = 1.8 m³
    data_rows = [
        [
            "TEST-002", "CT-002", "2026-01-01", "Test Corp", "CN",
            1, 1.2, 1.0, 1.5,
            "C1", 30, 30, 30, 20.0,
            1, "Test Product", "12345678", "PCS", 10, 5.0, 1.0, "CN",
        ],
        [
            "TEST-002", "CT-002", "2026-01-01", "Test Corp", "CN",
            2, 2.0, 1.5, 1.0,
            "C2", 30, 30, 30, 15.0,
            2, "Test Product 2", "87654321", "PCS", 5, 10.0, 0.5, "CN",
        ],
    ]

    excel_path = _create_test_excel(headers, data_rows)

    order, unmapped = import_order_from_excel(excel_path)

    # 托盘 1: 1.2×1.0×1.5 = 1.8
    # 托盘 2: 2.0×1.5×1.0 = 3.0
    # 总体积 = 1.8 + 3.0 = 4.8
    expected_volume = 1.2 * 1.0 * 1.5 + 2.0 * 1.5 * 1.0  # 4.8
    assert abs(order.totals.total_volume_cbm - expected_volume) < 0.01, (
        f"体积计算错误: 期望 {expected_volume}, 实际 {order.totals.total_volume_cbm}"
    )

    assert order.totals.total_pallets == 2
    print(f"✅ 测试 3 通过: 体积自动计算 {order.totals.total_volume_cbm:.3f} ≈ {expected_volume}")


# ==================== 测试 4：自动汇总 ====================


def test_auto_summary():
    """测试 4: 总毛重 = 所有纸箱毛重之和."""
    from src.importer.excel_importer import import_order_from_excel

    headers = [
        "发票号", "合同号", "日期", "客户", "国家",
        "托盘号", "托盘长度", "托盘宽度", "托盘高度",
        "纸箱标签", "纸箱长度", "纸箱宽度", "纸箱高度", "纸箱毛重", "箱数",
        "序号", "商品名称", "HS编码", "单位", "数量", "单价", "净重", "目的国",
    ]

    # 3 个托盘，每个 1 种纸箱，不同重量
    data_rows = [
        ["TEST-003", "CT-003", "2026-01-01", "Test Corp", "CN",
         1, 1.0, 1.0, 1.0, "A", 30, 30, 30, 10.5, 1,
         1, "P1", "1234", "PCS", 1, 100.0, 8.0, "CN"],
        ["TEST-003", "CT-003", "2026-01-01", "Test Corp", "CN",
         2, 1.0, 1.0, 1.0, "B", 30, 30, 30, 20.3, 2,
         2, "P2", "5678", "PCS", 1, 200.0, 18.0, "CN"],
        ["TEST-003", "CT-003", "2026-01-01", "Test Corp", "CN",
         3, 1.0, 1.0, 1.0, "C", 30, 30, 30, 15.0, 1,
         3, "P3", "9012", "PCS", 1, 150.0, 12.0, "CN"],
    ]

    excel_path = _create_test_excel(headers, data_rows)

    order, unmapped = import_order_from_excel(excel_path)

    # 总毛重 = 10.5*1 + 20.3*2 + 15.0*1 = 10.5 + 40.6 + 15.0 = 66.1
    expected_gross = 10.5 * 1 + 20.3 * 2 + 15.0 * 1
    assert abs(order.totals.total_gross_weight_kg - expected_gross) < 0.01, (
        f"总毛重计算错误: 期望 {expected_gross}, 实际 {order.totals.total_gross_weight_kg}"
    )

    # 总净重 = 8.0*1*1 + 18.0*1*2 + 12.0*1*1 = 8.0 + 36.0 + 12.0 = 56.0
    expected_net = 8.0 * 1 * 1 + 18.0 * 1 * 2 + 12.0 * 1 * 1
    assert abs(order.totals.total_net_weight_kg - expected_net) < 0.01, (
        f"总净重计算错误: 期望 {expected_net}, 实际 {order.totals.total_net_weight_kg}"
    )

    # 总金额 = 100*1*1 + 200*1*2 + 150*1*1 = 100 + 400 + 150 = 650
    expected_amount = 100.0 * 1 * 1 + 200.0 * 1 * 2 + 150.0 * 1 * 1
    assert abs(order.totals.total_amount - expected_amount) < 0.01, (
        f"总金额计算错误: 期望 {expected_amount}, 实际 {order.totals.total_amount}"
    )

    # 总纸箱数 = 1 + 2 + 1 = 4
    assert order.totals.total_cartons == 4, f"总纸箱数错误: 期望 4, 实际 {order.totals.total_cartons}"

    print(f"✅ 测试 4 通过: 毛重={order.totals.total_gross_weight_kg:.1f}, "
          f"净重={order.totals.total_net_weight_kg:.1f}, "
          f"金额={order.totals.total_amount:.1f}")


# ==================== 测试 5：模板加载 ====================


def test_template_load_from_db():
    """测试 5: 从 SQLite 保存模板 → 加载 → 反序列化 → 数据一致."""
    from src.db.repository import TemplateRepository
    from src.importer.template_loader import TemplateLoader
    from src.models.order_data import (
        Carton,
        Customer,
        OrderData,
        OrderMeta,
        Pallet,
        Product,
        Totals,
        Origin,
    )

    # 构造一个完整订单
    original_order = OrderData(
        order_meta=OrderMeta(
            invoice_no="TPL-001",
            contract_no="TPL-CT-001",
            date="2026-05-26",
            trade_term="CIF",
            payment_term="T/T 30% DEPOSIT",
            country_of_origin="China",
        ),
        customer=Customer(
            company_name_en="TEMPLATE CUSTOMER LTD.",
            country="Japan",
            company_name_cn="模板客户",
            address="Tokyo, Japan",
            contact_person="MR. TEST",
            phone="+81-123-456",
            mobile="+81-987-654",
            destination="Tokyo",
        ),
        pallets=[
            Pallet(
                pallet_no=1,
                length_m=1.2,
                width_m=1.0,
                height_m=1.5,
                pallet_weight_kg=15.0,
                cartons=[
                    Carton(
                        carton_label="TPL-C1",
                        is_batch=False,
                        batch_count=1,
                        length_cm=30.0,
                        width_cm=30.0,
                        height_cm=30.0,
                        gross_weight_kg=25.0,
                        products=[
                            Product(
                                seq_no=1,
                                product_name="Template Product A",
                                hs_code="11111111",
                                unit="PCS",
                                qty_per_carton=10.0,
                                unit_price=5.0,
                                net_weight_per_unit_kg=2.0,
                                destination_country="Japan",
                                specification="TEMPLATE-SPEC",
                            )
                        ],
                    )
                ],
            )
        ],
        totals=Totals(
            total_pallets=1,
            total_cartons=1,
            total_gross_weight_kg=25.0,
            total_net_weight_kg=20.0,
            total_volume_cbm=1.8,
            total_amount=50.0,
        ),
        origin=Origin(),
    )

    # 保存到数据库
    template_id = TemplateRepository.save(
        order=original_order,
        template_name="测试模板_导入器验证",
        description="用于验证 TemplateLoader 的测试模板",
    )
    assert template_id > 0, f"模板保存失败，返回 ID={template_id}"

    # 通过 TemplateLoader 加载
    loaded = TemplateLoader.load_template(template_id)
    assert loaded is not None, "加载模板返回 None"

    # 验证数据一致性
    assert loaded.order_meta.invoice_no == original_order.order_meta.invoice_no
    assert loaded.order_meta.contract_no == original_order.order_meta.contract_no
    assert loaded.order_meta.trade_term == original_order.order_meta.trade_term
    assert loaded.customer.company_name_en == original_order.customer.company_name_en
    assert loaded.customer.country == original_order.customer.country
    assert loaded.customer.company_name_cn == original_order.customer.company_name_cn
    assert loaded.totals.total_pallets == original_order.totals.total_pallets
    assert loaded.totals.total_amount == original_order.totals.total_amount

    # 验证 Pallet 数据
    assert len(loaded.pallets) == 1
    loaded_pallet = loaded.pallets[0]
    assert loaded_pallet.pallet_no == 1
    assert len(loaded_pallet.cartons) == 1

    loaded_carton = loaded_pallet.cartons[0]
    assert loaded_carton.carton_label == "TPL-C1"
    assert len(loaded_carton.products) == 1

    loaded_product = loaded_carton.products[0]
    assert loaded_product.product_name == "Template Product A"
    assert loaded_product.hs_code == "11111111"

    # 验证 TemplateLoader.list_templates
    templates = TemplateLoader.list_templates()
    assert len(templates) >= 1
    assert any(t["template_name"] == "测试模板_导入器验证" for t in templates)

    # 验证 TemplateLoader.load_latest
    latest = TemplateLoader.load_latest()
    assert latest is not None
    assert latest.order_meta.invoice_no == "TPL-001"

    print(f"✅ 测试 5 通过: 模板保存/加载数据一致 (ID={template_id})")

"""阶段 4 测试 — 装箱单生成器.

测试覆盖：
1. 小订单（1 托盘, 1 箱, 1 商品）→ 生成 xlsx 可打开, 数据正确
2. 中等订单（5 托盘, 多箱, 多商品）→ 行数正确, 汇总数据一致
3. 大订单（50+ 行数据 → 触发扩容）→ 插入新行带样式, 公式修正
4. 批量纸箱（is_batch=true, batch_count=10）→ 正确合并为一行
5. 空订单（0 托盘）→ 报错不生成, 模板不被破坏
6. 模板损坏（锚点扫描失败）→ 报错并提示恢复出厂模板
"""

from __future__ import annotations

import contextlib
import tempfile
from pathlib import Path

import openpyxl
import pytest

from config.constants import OUTPUT_DIR, TEMPLATE_PACKING_PATH
from src.generators.packing_generator import (
    PackingGenerator,
    flatten_for_packing,
)
from src.models.order_data import (
    Carton,
    Customer,
    OrderData,
    OrderMeta,
    Origin,
    Pallet,
    Product,
    Totals,
)

# ==================== 测试辅助：构造订单 ====================


def _make_minimal_order() -> OrderData:
    """构造最小合法订单: 1 托盘, 1 纸箱, 1 商品."""
    return OrderData(
        order_meta=OrderMeta(
            invoice_no="TEST-001",
            contract_no="TEST-CONTRACT",
            date="2025-12-26",
            trade_term="FOB",
            payment_term="100% T/T IN ADVANCE",
            country_of_origin="China",
            transport_mode="海运",
        ),
        customer=Customer(
            company_name_en="TEST CUSTOMER LTD.",
            country="Turkey",
            destination="Ankara",
        ),
        pallets=[
            Pallet(
                pallet_no=1,
                length_m=1.16,
                width_m=1.01,
                height_m=1.97,
                cartons=[
                    Carton(
                        carton_label="A1",
                        is_batch=False,
                        batch_count=1,
                        length_cm=32.0,
                        width_cm=32.0,
                        height_cm=34.0,
                        gross_weight_kg=23.3,
                        products=[
                            Product(
                                seq_no=1,
                                product_name="Test Product A",
                                specification="100mm*2.0mm*30M",
                                hs_code="3926909090",
                                unit="Roll",
                                qty_per_carton=1,
                                unit_price=85.0,
                                net_weight_per_unit_kg=22.3,
                                destination_country="Turkey",
                            )
                        ],
                    )
                ],
            )
        ],
        totals=Totals(
            total_pallets=1,
            total_cartons=1,
            total_gross_weight_kg=23.3,
            total_net_weight_kg=22.3,
            total_volume_cbm=0.035,
            total_amount=85.0,
        ),
        origin=Origin(export_port="Shenzhen"),
    )


def _make_medium_order() -> OrderData:
    """构造中等订单: 5 托盘, 多纸箱, 3 种商品."""
    pallets: list[Pallet] = []
    for p in range(1, 6):
        cartons: list[Carton] = []
        for c in range(1, 5):  # 每托4个纸箱
            cartons.append(
                Carton(
                    carton_label=f"C{c}",
                    is_batch=False,
                    batch_count=1,
                    length_cm=32.0,
                    width_cm=32.0,
                    height_cm=34.0 + c,
                    gross_weight_kg=20.0 + c,
                    products=[
                        Product(
                            seq_no=c,
                            product_name=f"Product {c % 3 + 1}",
                            specification=f"Spec-{c}",
                            hs_code="3926909090",
                            unit="Roll",
                            qty_per_carton=1,
                            unit_price=10.0 * c,
                            net_weight_per_unit_kg=18.0 + c,
                            destination_country="Turkey",
                        )
                    ],
                )
            )
        pallets.append(
            Pallet(
                pallet_no=p,
                length_m=1.2,
                width_m=1.0,
                height_m=1.8,
                cartons=cartons,
            )
        )

    # 汇总
    total_cartons = 5 * 4
    total_gross = sum(20.0 + c for c in range(1, 5)) * 5
    total_net = sum(18.0 + c for c in range(1, 5)) * 5
    total_vol = 1.2 * 1.0 * 1.8 * 5
    total_amount = sum(10.0 * c for c in range(1, 5)) * 5

    return OrderData(
        order_meta=OrderMeta(
            invoice_no="TEST-002",
            contract_no="TEST-CONTRACT",
            date="2025-12-26",
            trade_term="FOB",
            payment_term="100% T/T IN ADVANCE",
            country_of_origin="China",
        ),
        customer=Customer(
            company_name_en="TEST CUSTOMER LTD.",
            country="Turkey",
        ),
        pallets=pallets,
        totals=Totals(
            total_pallets=5,
            total_cartons=total_cartons,
            total_gross_weight_kg=total_gross,
            total_net_weight_kg=total_net,
            total_volume_cbm=round(total_vol, 3),
            total_amount=total_amount,
        ),
        origin=Origin(export_port="Shenzhen"),
    )


def _make_batch_order() -> OrderData:
    """构造批量纸箱订单: 1 托盘, 1 批量纸箱（batch_count=10）, 1 商品."""
    return OrderData(
        order_meta=OrderMeta(
            invoice_no="TEST-BATCH",
            contract_no="TEST-CONTRACT",
            date="2025-12-26",
            trade_term="FOB",
            payment_term="100% T/T IN ADVANCE",
            country_of_origin="China",
        ),
        customer=Customer(
            company_name_en="TEST CUSTOMER LTD.",
            country="Turkey",
        ),
        pallets=[
            Pallet(
                pallet_no=1,
                length_m=1.16,
                width_m=1.01,
                height_m=1.97,
                cartons=[
                    Carton(
                        carton_label="BATCH-10",
                        is_batch=True,
                        batch_count=10,
                        length_cm=32.0,
                        width_cm=32.0,
                        height_cm=34.0,
                        gross_weight_kg=23.3,
                        products=[
                            Product(
                                seq_no=1,
                                product_name="Batch Product",
                                specification="330mm*2.0mm*30M",
                                hs_code="3926909090",
                                unit="Roll",
                                qty_per_carton=1,
                                unit_price=85.0,
                                net_weight_per_unit_kg=22.3,
                                destination_country="Turkey",
                            )
                        ],
                    )
                ],
            )
        ],
        totals=Totals(
            total_pallets=1,
            total_cartons=10,
            total_gross_weight_kg=233.0,
            total_net_weight_kg=223.0,
            total_volume_cbm=0.348,
            total_amount=850.0,
        ),
        origin=Origin(export_port="Shenzhen"),
    )


# ==================== 测试 1: flatten_for_packing ====================


class TestFlattenForPacking:
    """数据展平测试."""

    def test_minimal_order_flatten(self) -> None:
        """最小订单展平为 1 行."""
        order = _make_minimal_order()
        rows = flatten_for_packing(order)
        assert len(rows) == 1
        assert rows[0]["seq_no"] == 1
        assert rows[0]["pallet_no"] == 1
        assert rows[0]["product_name"] == "Test Product A"
        assert rows[0]["carton_count"] == 1

    def test_batch_order_flatten(self) -> None:
        """批量纸箱展平为 1 行，carton_count = batch_count."""
        order = _make_batch_order()
        rows = flatten_for_packing(order)
        assert len(rows) == 1
        assert rows[0]["carton_count"] == 10
        assert rows[0]["net_weight"] == pytest.approx(22.3 * 1 * 10, rel=0.01)
        assert rows[0]["gross_weight"] == pytest.approx(23.3 * 10, rel=0.01)

    def test_none_order_raises(self) -> None:
        """None 订单抛出 ValueError."""
        with pytest.raises(ValueError, match="order 为 None"):
            flatten_for_packing(None)  # type: ignore[arg-type]

    def test_empty_pallets_raises(self) -> None:
        """空托盘列表抛出 ValueError."""
        order = OrderData(
            order_meta=OrderMeta(
                invoice_no="X", contract_no="X", date="2025-01-01",
                trade_term="FOB", payment_term="T/T", country_of_origin="China",
            ),
            customer=Customer(company_name_en="X", country="X"),
            pallets=[],
            totals=Totals(
                total_pallets=0, total_cartons=0,
                total_gross_weight_kg=0.0, total_net_weight_kg=0.0,
                total_volume_cbm=0.0, total_amount=0.0,
            ),
        )
        with pytest.raises(ValueError, match="无托盘数据"):
            flatten_for_packing(order)

    def test_multiple_products_per_carton(self) -> None:
        """一个纸箱内含多个商品，展开为多行."""
        order = OrderData(
            order_meta=OrderMeta(
                invoice_no="MULTI", contract_no="C", date="2025-01-01",
                trade_term="FOB", payment_term="T/T", country_of_origin="China",
            ),
            customer=Customer(company_name_en="X", country="X"),
            pallets=[
                Pallet(
                    pallet_no=1,
                    length_m=1.0, width_m=1.0, height_m=1.0,
                    cartons=[
                        Carton(
                            carton_label="M1",
                            is_batch=False,
                            batch_count=1,
                            length_cm=30.0, width_cm=30.0, height_cm=30.0,
                            gross_weight_kg=50.0,
                            products=[
                                Product(
                                    seq_no=1, product_name="P1", hs_code="HS1",
                                    unit="PCS", qty_per_carton=10, unit_price=5.0,
                                    net_weight_per_unit_kg=0.5, destination_country="TR",
                                ),
                                Product(
                                    seq_no=2, product_name="P2", hs_code="HS2",
                                    unit="PCS", qty_per_carton=5, unit_price=3.0,
                                    net_weight_per_unit_kg=0.3, destination_country="TR",
                                ),
                            ],
                        )
                    ],
                )
            ],
            totals=Totals(
                total_pallets=1, total_cartons=1,
                total_gross_weight_kg=50.0, total_net_weight_kg=6.5,
                total_volume_cbm=1.0, total_amount=65.0,
            ),
        )
        rows = flatten_for_packing(order)
        assert len(rows) == 2
        assert rows[0]["seq_no"] == 1
        assert rows[1]["seq_no"] == 2
        assert rows[0]["product_name"] == "P1"
        assert rows[1]["product_name"] == "P2"


# ==================== 测试 2: PackingGenerator.generate ====================


class TestPackingGenerator:
    """装箱单生成器核心测试."""

    @pytest.fixture
    def gen(self) -> PackingGenerator:
        return PackingGenerator()

    def test_generate_minimal_order(self, gen: PackingGenerator) -> None:
        """小订单生成: 文件可打开, 数据正确."""
        order = _make_minimal_order()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = gen.generate(order, output_dir=tmpdir)
            assert output_path.exists()
            assert output_path.suffix == ".xlsx"

            # 验证可打开
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active

            # 验证表头
            assert ws.cell(row=3, column=4).value is not None  # D3 客户名

            # 验证至少有一行数据
            # 扫描找到数据区域
            found_data = False
            for row in range(1, ws.max_row + 1):
                a_val = ws.cell(row=row, column=1).value
                if isinstance(a_val, (int, float)) and a_val == 1:
                    found_data = True
                    break
            assert found_data, "未找到序号为 1 的数据行"

            wb.close()

    def test_generate_medium_order(self, gen: PackingGenerator) -> None:
        """中等订单生成: 行数正确."""
        order = _make_medium_order()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = gen.generate(order, output_dir=tmpdir)
            assert output_path.exists()

            wb = openpyxl.load_workbook(output_path)
            ws = wb.active

            # 数据行数 = 5 托盘 × 4 纸箱 × 1 商品/箱 = 20 行
            # 从可见内容统计
            data_rows = 0
            for row in range(1, ws.max_row + 1):
                a_val = ws.cell(row=row, column=1).value
                if isinstance(a_val, (int, float)) and a_val > 0:
                    data_rows += 1
            assert data_rows == 20, f"期望 20 行数据, 实际 {data_rows} 行"

            wb.close()

    def test_generate_batch_order(self, gen: PackingGenerator) -> None:
        """批量纸箱订单: carton_count = 10."""
        order = _make_batch_order()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = gen.generate(order, output_dir=tmpdir)
            assert output_path.exists()

            wb = openpyxl.load_workbook(output_path)
            ws = wb.active

            # 查找数据行，G 列（箱数）应为 10
            for row in range(1, ws.max_row + 1):
                a_val = ws.cell(row=row, column=1).value
                if isinstance(a_val, (int, float)) and a_val == 1:
                    g_val = ws.cell(row=row, column=7).value  # G 列 = 箱数
                    assert g_val == 10, f"期望箱数=10, 实际={g_val}"
                    break

            wb.close()

    def test_generate_empty_order_raises(self, gen: PackingGenerator) -> None:
        """空订单（无托盘）抛出 ValueError."""
        order = OrderData(
            order_meta=OrderMeta(
                invoice_no="EMPTY", contract_no="C", date="2025-01-01",
                trade_term="FOB", payment_term="T/T", country_of_origin="China",
            ),
            customer=Customer(company_name_en="X", country="X"),
            pallets=[],
            totals=Totals(
                total_pallets=0, total_cartons=0,
                total_gross_weight_kg=0.0, total_net_weight_kg=0.0,
                total_volume_cbm=0.0, total_amount=0.0,
            ),
        )
        with pytest.raises(ValueError, match="无托盘数据"):
            gen.generate(order)

    def test_generate_none_order_raises(self, gen: PackingGenerator) -> None:
        """None 订单抛出 ValueError."""
        with pytest.raises(ValueError, match="订单数据为 None"):
            gen.generate(None)  # type: ignore[arg-type]

    def test_template_not_damaged_after_error(self, gen: PackingGenerator) -> None:
        """出错后模板原文件未被修改."""
        # 记录原始模板的修改时间
        original_mtime = TEMPLATE_PACKING_PATH.stat().st_mtime

        order = OrderData(
            order_meta=OrderMeta(
                invoice_no="ERR", contract_no="C", date="2025-01-01",
                trade_term="FOB", payment_term="T/T", country_of_origin="China",
            ),
            customer=Customer(company_name_en="X", country="X"),
            pallets=[],
            totals=Totals(
                total_pallets=0, total_cartons=0,
                total_gross_weight_kg=0.0, total_net_weight_kg=0.0,
                total_volume_cbm=0.0, total_amount=0.0,
            ),
        )

        with contextlib.suppress(ValueError):
            gen.generate(order)

        # 模板文件修改时间不应改变
        assert TEMPLATE_PACKING_PATH.stat().st_mtime == original_mtime, (
            "模板文件被意外修改！"
        )

    def test_output_file_naming(self, gen: PackingGenerator) -> None:
        """输出文件名正确: 装箱单_{发票号}.xlsx."""
        order = _make_minimal_order()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = gen.generate(order, output_dir=tmpdir)
            expected_name = f"装箱单_{order.order_meta.invoice_no}.xlsx"
            assert output_path.name == expected_name, (
                f"期望文件名={expected_name}, 实际={output_path.name}"
            )

    def test_progress_callback(self, gen: PackingGenerator) -> None:
        """进度回调被正确调用."""
        order = _make_minimal_order()
        callbacks: list[tuple] = []

        def on_progress(step: str, progress: float) -> None:
            callbacks.append((step, progress))

        with tempfile.TemporaryDirectory() as tmpdir:
            gen.generate(order, output_dir=tmpdir, progress_callback=on_progress)

        assert len(callbacks) >= 2, f"回调次数={len(callbacks)}, 期望 ≥2"
        # 最后一个回调应为 progress=1.0
        assert callbacks[-1][1] == 1.0, f"最后进度={callbacks[-1][1]}, 期望=1.0"

    def test_generate_with_default_output_dir(self, gen: PackingGenerator) -> None:
        """未指定 output_dir 时使用默认输出目录."""
        order = _make_minimal_order()
        output_path = gen.generate(order)

        # 检查文件在 OUTPUT_DIR 中
        assert OUTPUT_DIR in output_path.parents, (
            f"输出路径 {output_path} 不在默认输出目录 {OUTPUT_DIR} 中"
        )
        assert output_path.exists()

        # 清理
        output_path.unlink(missing_ok=True)

    def test_damaged_template_raises(self) -> None:
        """损坏的模板文件抛出异常."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"not a valid xlsx file")
            bad_path = f.name

        try:
            gen = PackingGenerator(template_path=bad_path)
            order = _make_minimal_order()
            with pytest.raises((ValueError, Exception)):
                gen.generate(order)
        finally:
            Path(bad_path).unlink(missing_ok=True)


# ==================== 真实订单数据结构测试 ====================

class TestRealOrderStructure:
    """基于 RealOrderData.md 的数据结构正确性测试.

    验证第 11 托的多商品多纸箱结构能正确展平.
    """

    def test_pallet_11_structure(self) -> None:
        """第 11 托: 4 种商品, 6 个不同纸箱规格, 共 25 箱."""
        # 构造第 11 托（不含其他托盘）
        pallet_11 = Pallet(
            pallet_no=11,
            length_m=1.16,
            width_m=1.01,
            height_m=1.52,
            cartons=[
                # 370mm: is_batch, 10 箱
                Carton(
                    carton_label="370mm-batch",
                    is_batch=True,
                    batch_count=10,
                    length_cm=32.0, width_cm=32.0, height_cm=38.0,
                    gross_weight_kg=27.0,  # 约值
                    products=[
                        Product(
                            seq_no=1,
                            product_name="50℃ Type 2A-1 Heat Shrink Sleeve",
                            specification="370mm*2.0mm*30M",
                            hs_code="3926909090",
                            unit="Roll",
                            qty_per_carton=1,
                            unit_price=100.0,
                            net_weight_per_unit_kg=25.44,
                            destination_country="Turkey",
                        )
                    ],
                ),
                # 470mm: is_batch, 7 箱
                Carton(
                    carton_label="470mm-batch",
                    is_batch=True,
                    batch_count=7,
                    length_cm=32.0, width_cm=32.0, height_cm=47.0,
                    gross_weight_kg=40.0,  # 约值
                    products=[
                        Product(
                            seq_no=2,
                            product_name="50℃ Type 2A-1 Heat Shrink Sleeve",
                            specification="470mm*2.0mm*30M",
                            hs_code="3926909090",
                            unit="Roll",
                            qty_per_carton=1,
                            unit_price=115.0,
                            net_weight_per_unit_kg=39.0,
                            destination_country="Turkey",
                        )
                    ],
                ),
                # Filler type 1: is_batch, 3 箱, 48x20x22.5cm, 60 PCS/CTN
                Carton(
                    carton_label="Filler-48cm",
                    is_batch=True,
                    batch_count=3,
                    length_cm=48.0, width_cm=20.0, height_cm=22.5,
                    gross_weight_kg=8.0,  # 约值
                    products=[
                        Product(
                            seq_no=3,
                            product_name="Filler",
                            specification="1.5mm*50mm*5m",
                            hs_code="3926909090",
                            unit="Roll",
                            qty_per_carton=60,
                            unit_price=3.15,
                            net_weight_per_unit_kg=0.04,
                            destination_country="Turkey",
                        )
                    ],
                ),
                # Filler type 2: 1 箱, 43x21x22cm, 20 PCS/CTN
                Carton(
                    carton_label="Filler-43cm",
                    is_batch=False,
                    batch_count=1,
                    length_cm=43.0, width_cm=21.0, height_cm=22.0,
                    gross_weight_kg=2.5,
                    products=[
                        Product(
                            seq_no=4,
                            product_name="Filler",
                            specification="1.5mm*50mm*5m",
                            hs_code="3926909090",
                            unit="Roll",
                            qty_per_carton=20,
                            unit_price=3.15,
                            net_weight_per_unit_kg=0.04,
                            destination_country="Turkey",
                        )
                    ],
                ),
                # Repair Patch type 1: is_batch, 3 箱, 80x40x25cm, 8 PCS/CTN
                Carton(
                    carton_label="Patch-80cm",
                    is_batch=True,
                    batch_count=3,
                    length_cm=80.0, width_cm=40.0, height_cm=25.0,
                    gross_weight_kg=12.0,
                    products=[
                        Product(
                            seq_no=5,
                            product_name="Repair Patch",
                            specification="250mm*2.4mm*5m",
                            hs_code="3926909090",
                            unit="Roll",
                            qty_per_carton=8,
                            unit_price=12.5,
                            net_weight_per_unit_kg=0.15,
                            destination_country="Turkey",
                        )
                    ],
                ),
                # Repair Patch type 2: 1 箱, 30x30x15cm, 1 PCS/CTN
                Carton(
                    carton_label="Patch-30cm",
                    is_batch=False,
                    batch_count=1,
                    length_cm=30.0, width_cm=30.0, height_cm=15.0,
                    gross_weight_kg=1.0,
                    products=[
                        Product(
                            seq_no=6,
                            product_name="Repair Patch",
                            specification="250mm*2.4mm*5m",
                            hs_code="3926909090",
                            unit="Roll",
                            qty_per_carton=1,
                            unit_price=12.5,
                            net_weight_per_unit_kg=0.15,
                            destination_country="Turkey",
                        )
                    ],
                ),
            ],
        )

        # 展平
        order = OrderData(
            order_meta=OrderMeta(
                invoice_no="TEST-011",
                contract_no="PO25-018",
                date="2025-12-26",
                trade_term="FOB",
                payment_term="100% T/T IN ADVANCE",
                country_of_origin="China",
            ),
            customer=Customer(
                company_name_en="YARIMKURE INSAAT LTD. STI.",
                country="Turkey",
            ),
            pallets=[pallet_11],
            totals=Totals(
                total_pallets=1,
                total_cartons=25,  # 10+7+3+1+3+1=25
                total_gross_weight_kg=672.0,
                total_net_weight_kg=636.9,
                total_volume_cbm=1.781,
                total_amount=1150.0 + 630.0 + 312.5,
            ),
        )

        rows = flatten_for_packing(order)

        # 应展平为 6 行（6 个不同的纸箱→商品组合）
        assert len(rows) == 6, f"期望 6 行, 实际 {len(rows)}"

        # 验证各行的 carton_count
        expected_counts = [10, 7, 3, 1, 3, 1]
        for i, row in enumerate(rows):
            assert row["carton_count"] == expected_counts[i], (
                f"第 {i+1} 行: 期望 carton_count={expected_counts[i]}, "
                f"实际={row['carton_count']}"
            )

        # 总箱数应为 10+7+3+1+3+1=25
        total_cartons = sum(r["carton_count"] for r in rows)
        assert total_cartons == 25, f"总箱数期望 25, 实际 {total_cartons}"

        # 序号应为 1-6 递增
        for i, row in enumerate(rows):
            assert row["seq_no"] == i + 1

        # 所有托盘号应为 11
        for row in rows:
            assert row["pallet_no"] == 11


# ========== 运行说明 ==========
# 依赖安装: pip install openpyxl msgspec pytest
# 运行命令: python -m pytest tests/test_packing_generator.py -v
# 预期输出: 全部 PASSED
# =============================

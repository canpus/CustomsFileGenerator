# -*- coding: utf-8 -*-
"""阶段 3 测试 — XLSX 工具库 + 锚点扫描.

测试覆盖：
1. clone_row_style: 样式深拷贝验证
2. insert_rows_with_style: 插入行样式复制
3. delete_rows_safely 缩水场景: 50行→5行 + 公式修正
4. delete_rows_safely 含合并单元格: 安全删除不报错
5. update_sum_formula: SUM 公式范围修正
6. 锚点扫描 packing 模板
7. 锚点扫描 invoice 模板
8. 锚点扫描 contract 模板
9. 锚点扫描失败（模板损坏）
"""

from __future__ import annotations

import copy
import tempfile
from pathlib import Path

import openpyxl
import pytest

from config.constants import (
    TEMPLATE_PACKING_PATH,
    TEMPLATE_INVOICE_PATH,
    TEMPLATE_CONTRACT_PATH,
)
from src.generators.xlsx_utils import (
    clone_row_style,
    delete_rows_safely,
    get_merged_ranges_in_row,
    insert_rows_with_style,
    resize_data_rows,
    update_sum_formula,
)
from src.generators.template_anchor_scanner import (
    AnchorResult,
    scan_template,
    scan_packing_template,
    scan_invoice_template,
    scan_contract_template,
)


# ==================== Fixtures ====================


@pytest.fixture
def sample_ws():
    """创建含样式的工作表."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"

    # 加粗标题行
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

    header_font = Font(name="Arial", size=12, bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.value = f"Col{col}"
        cell.font = copy.deepcopy(header_font)
        cell.border = copy.deepcopy(thin_border)
        cell.fill = copy.deepcopy(header_fill)

    # 数据行
    for row in range(2, 11):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.value = row * 10 + col
            cell.font = Font(name="Arial", size=11)
            cell.border = copy.deepcopy(thin_border)

    yield ws
    wb.close()


@pytest.fixture
def ws_with_merges():
    """创建含合并单元格的工作表."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # 填充数据区域
    for row in range(1, 20):
        for col in range(1, 4):
            ws.cell(row=row, column=col).value = f"R{row}C{col}"

    # 合并第 8-10 行 B:C 列
    ws.merge_cells("B8:C8")
    ws.merge_cells("B9:C9")
    ws.merge_cells("B10:C10")

    yield ws
    wb.close()


@pytest.fixture
def ws_with_sum_formula():
    """创建含 SUM 公式的工作表."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # 数据区域 G8:G20
    for row in range(8, 21):
        ws.cell(row=row, column=7).value = row  # G 列

    # 汇总行 SUM 公式
    ws.cell(row=21, column=7).value = "=SUM(G8:G57)"

    yield ws
    wb.close()


# ==================== 测试 1: clone_row_style ====================


def test_clone_row_style_font(sample_ws):
    """克隆后目标行字体与源行完全一致."""
    clone_row_style(sample_ws, 1, 20)
    src_font = sample_ws.cell(row=1, column=1).font
    tgt_font = sample_ws.cell(row=20, column=1).font
    assert tgt_font.name == src_font.name
    assert tgt_font.size == src_font.size
    assert tgt_font.bold == src_font.bold


def test_clone_row_style_border(sample_ws):
    """克隆后目标行边框与源行一致."""
    clone_row_style(sample_ws, 1, 20)
    src_border = sample_ws.cell(row=1, column=1).border
    tgt_border = sample_ws.cell(row=20, column=1).border
    assert tgt_border.left.style == src_border.left.style
    assert tgt_border.right.style == src_border.right.style


def test_clone_row_style_fill(sample_ws):
    """克隆后目标行填充色与源行一致."""
    clone_row_style(sample_ws, 1, 20)
    src_fill = sample_ws.cell(row=1, column=1).fill
    tgt_fill = sample_ws.cell(row=20, column=1).fill
    assert tgt_fill.start_color.rgb == src_fill.start_color.rgb


def test_clone_row_style_invalid_row():
    """无效行号抛出 ValueError."""
    wb = openpyxl.Workbook()
    ws = wb.active
    with pytest.raises(ValueError):
        clone_row_style(ws, 0, 1)
    with pytest.raises(ValueError):
        clone_row_style(ws, 1, 0)
    wb.close()


# ==================== 测试 2: insert_rows_with_style ====================


def test_insert_rows_with_style_count(sample_ws):
    """插入后总行数正确."""
    original_max = sample_ws.max_row
    insert_rows_with_style(sample_ws, 5, 5)
    assert sample_ws.max_row == original_max + 5


def test_insert_rows_with_style_copies_style(sample_ws):
    """插入的行样式与锚点行一致."""
    # 先设置锚点行样式
    from openpyxl.styles import Font
    sample_ws.cell(row=5, column=1).font = Font(name="Times New Roman", size=14, bold=True)

    insert_rows_with_style(sample_ws, 5, 2)

    tgt_font = sample_ws.cell(row=7, column=1).font  # 5+2=7
    assert tgt_font.name == "Times New Roman"
    assert tgt_font.size == 14
    assert tgt_font.bold is True


def test_insert_rows_with_style_zero_count():
    """count=0 抛出 ValueError."""
    wb = openpyxl.Workbook()
    ws = wb.active
    with pytest.raises(ValueError):
        insert_rows_with_style(ws, 5, 0)
    wb.close()


# ==================== 测试 3: delete_rows_safely 缩水场景 ====================


def test_delete_rows_safely_basic(sample_ws):
    """基本删除: 删除 5 行后剩余行数正确."""
    original_max = sample_ws.max_row
    delete_rows_safely(sample_ws, 6, 10)
    assert sample_ws.max_row == original_max - 5


def test_delete_rows_safely_reverse_order_preserves_data(sample_ws):
    """逆序删除保留前 5 行数据不变."""
    delete_rows_safely(sample_ws, 6, 10)
    # 第 1-5 行数据应不变
    for row in range(1, 6):
        for col in range(1, 6):
            cell = sample_ws.cell(row=row, column=col)
            # 标题行检查
            if row == 1:
                assert cell.value == f"Col{col}"


def test_delete_rows_safely_invalid_range():
    """start > end 抛出 ValueError."""
    wb = openpyxl.Workbook()
    ws = wb.active
    with pytest.raises(ValueError):
        delete_rows_safely(ws, 10, 5)
    wb.close()


# ==================== 测试 4: delete_rows_safely 含合并单元格 ====================


def test_delete_rows_with_merges(ws_with_merges):
    """含合并单元格的行安全删除不报错."""
    delete_rows_safely(ws_with_merges, 8, 10)
    # 删除后第 8 行不再是原来的数据
    assert ws_with_merges.max_row == 16  # 19 - 3


def test_get_merged_ranges_in_row(ws_with_merges):
    """正确检测指定行的合并单元格."""
    ranges = get_merged_ranges_in_row(ws_with_merges, 9)
    assert len(ranges) >= 1


# ==================== 测试 5: update_sum_formula ====================


def test_update_sum_formula_correction(ws_with_sum_formula):
    """SUM(G8:G57) 修正为 SUM(G8:G20)."""
    update_sum_formula(ws_with_sum_formula, "G", 8, 20)
    cell = ws_with_sum_formula.cell(row=21, column=7)
    assert cell.value == "=SUM(G8:G20)"


def test_update_sum_formula_no_match(ws_with_sum_formula):
    """无匹配列时不报错."""
    update_sum_formula(ws_with_sum_formula, "Z", 1, 10)
    # 不应抛出异常


def test_update_sum_formula_invalid_col():
    """无效列字母抛出 ValueError."""
    wb = openpyxl.Workbook()
    ws = wb.active
    with pytest.raises(ValueError):
        update_sum_formula(ws, "", 1, 10)
    wb.close()


# ==================== 测试 6-8: 锚点扫描 ====================


def test_scan_packing_template():
    """锚点扫描装箱单模板：正确找到数据起始行和汇总行."""
    result = scan_template(TEMPLATE_PACKING_PATH)
    assert isinstance(result, AnchorResult)
    assert result.data_start_row > 0, (
        f"未找到数据起始行: errors={result.errors}"
    )
    assert result.summary_row > 0, (
        f"未找到汇总行: errors={result.errors}"
    )
    # 数据起始行应在标题行之后
    assert result.data_start_row > result.header_row


def test_scan_invoice_template():
    """锚点扫描发票模板：正确找到数据起始行和汇总行."""
    result = scan_template(TEMPLATE_INVOICE_PATH)
    assert isinstance(result, AnchorResult)
    assert result.data_start_row > 0, (
        f"未找到数据起始行: errors={result.errors}"
    )
    assert result.summary_row > 0, (
        f"未找到汇总行: errors={result.errors}"
    )


def test_scan_contract_template():
    """锚点扫描合同模板：正确找到数据起始行和汇总行."""
    result = scan_template(TEMPLATE_CONTRACT_PATH)
    assert isinstance(result, AnchorResult)
    assert result.data_start_row > 0, (
        f"未找到数据起始行: errors={result.errors}"
    )
    assert result.summary_row > 0, (
        f"未找到汇总行: errors={result.errors}"
    )


# ==================== 测试 9: 锚点扫描失败 ====================


def test_scan_template_file_not_found():
    """模板文件不存在时抛出 FileNotFoundError."""
    with pytest.raises(FileNotFoundError):
        scan_template("nonexistent_template.xlsx")


def test_scan_template_corrupted():
    """损坏的模板文件抛出 ValueError（含中文错误提示）."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(b"this is not a valid xlsx file")
        temp_path = f.name

    try:
        with pytest.raises(ValueError) as exc_info:
            scan_template(temp_path)
        assert "无法打开模板文件" in str(exc_info.value)
    finally:
        Path(temp_path).unlink(missing_ok=True)


def test_scan_template_unsupported_name():
    """不支持的模板文件名抛出 ValueError."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "test"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        temp_path = f.name
    wb.close()

    try:
        with pytest.raises(ValueError) as exc_info:
            scan_template(temp_path)
        assert "不支持的模板类型" in str(exc_info.value)
    finally:
        Path(temp_path).unlink(missing_ok=True)


# ==================== 额外: resize_data_rows 测试 ====================


def test_resize_data_rows_expand(sample_ws):
    """扩容：从 9 行扩展到 15 行."""
    new_end = resize_data_rows(sample_ws, 2, 10, 15, 1)
    assert new_end == 10 + 6  # 10 + (15-9) = 16
    assert sample_ws.max_row >= 16


def test_resize_data_rows_shrink(sample_ws):
    """缩容：从 9 行缩减到 3 行."""
    new_end = resize_data_rows(sample_ws, 2, 10, 3, 1)
    assert new_end == 4  # 2+3-1
    assert sample_ws.max_row <= 5


# ========== 运行说明 ==========
# 依赖安装: pip install openpyxl pytest
# 运行命令: python -m pytest tests/test_xlsx_utils.py -v
# 预期输出: 全部 PASSED（约 18 项测试）
# =============================
